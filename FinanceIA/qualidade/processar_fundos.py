"""
processar_fundos.py
Controle de qualidade dos dados qualitativos de fundos.

Etapa 1 (Sonnet): Mescla duplicatas (mesmo CNPJ em múltiplas plataformas)
Etapa 2 (Haiku):  Valida cada fundo contra as diretrizes da subcategoria

Output: fundos_revisados.xlsx
  - Aba "Fundos Revisados": 1 linha por CNPJ único + validacao_status + validacao_obs
  - Aba "Disponibilidade": CNPJ → plataforma (grade de checkmarks)

Uso:
    python processar_fundos.py
    python processar_fundos.py --input meu_arquivo.xlsx
"""

import argparse
import json
import time
from datetime import datetime

import anthropic
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# ─── CONFIGURAÇÃO ────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = "COLE_SUA_CHAVE_AQUI"   # sk-ant-api03-...

EXCEL_INPUT    = "consolidado_fundos_total.xlsx"
EXCEL_OUTPUT   = "fundos_revisados.xlsx"
CHECKPOINT_MESCLA   = "progress_mescla.json"
CHECKPOINT_VALIDACAO = "progress_validacao.json"

PLATAFORMAS = ["XP", "BTG", "Itaú", "Bradesco", "Santander", "Inter"]
PRIORIDADE_PLAT = ["XP", "BTG", "Itaú", "Bradesco", "Santander", "Inter"]

QUALITATIVE_COLS = [
    "quando_indicar", "quando_nao_indicar",
    "vantagens", "desvantagens", "alertas",
    "descricao_simples", "descricao_tecnica",
]

MODEL_SONNET = "claude-sonnet-4-20250514"
MODEL_HAIKU  = "claude-haiku-4-5-20251001"

# ─── DIRETRIZES POR SUBCATEGORIA ─────────────────────────────────────────────
# Diretrizes mínimas esperadas para cada subcategoria — usadas na validação

DIRETRIZES = {
    # ── RENDA FIXA ────────────────────────────────────────────────────────
    "Soberano": {
        "quando_indicar": "Reserva de emergência, perfil conservador, curto prazo, liquidez imediata.",
        "quando_nao_indicar": "Investidores que buscam rentabilidade acima do CDI.",
        "alertas": "Risco de retorno abaixo da inflação em cenários de juros baixos.",
    },
    "Crédito Privado": {
        "quando_indicar": "Perfil moderado, médio prazo, busca de prêmio sobre CDI.",
        "quando_nao_indicar": "Perfil conservador ou quem necessita de liquidez imediata.",
        "alertas": "Risco de crédito. Liquidez restrita. Come-cotas semestral.",
    },
    "Inflação": {
        "quando_indicar": "Proteção contra inflação, perfil conservador a moderado, longo prazo.",
        "quando_nao_indicar": "Curto prazo. Cenários de queda de inflação/juros reais.",
        "alertas": "Marcação a mercado pode gerar volatilidade no curto prazo.",
    },
    "Prefixado": {
        "quando_indicar": "Cenário de queda de juros, perfil moderado, médio a longo prazo.",
        "quando_nao_indicar": "Cenário de alta de juros. Perfil conservador com baixa tolerância a volatilidade.",
        "alertas": "Risco de mercado elevado em cenários de alta de juros.",
    },
    # ── MULTIMERCADO ──────────────────────────────────────────────────────
    "Macro": {
        "quando_indicar": "Perfil moderado a arrojado, médio a longo prazo, diversificação da carteira.",
        "quando_nao_indicar": "Perfil conservador ou curto prazo.",
        "alertas": "Volatilidade maior que renda fixa. Resultados podem ser negativos no curto prazo.",
    },
    "Long & Short": {
        "quando_indicar": "Perfil arrojado, descorrelação com bolsa, médio a longo prazo.",
        "quando_nao_indicar": "Perfil conservador. Não recomendado como único investimento.",
        "alertas": "Estratégia complexa. Pode ter resultado negativo mesmo com bolsa positiva.",
    },
    "Quantitativo": {
        "quando_indicar": "Perfil moderado a arrojado, descorrelação, diversificação.",
        "quando_nao_indicar": "Perfil conservador.",
        "alertas": "Performance pode divergir significativamente em eventos de mercado atípicos.",
    },
    # ── RENDA VARIÁVEL ────────────────────────────────────────────────────
    "Long Only": {
        "quando_indicar": "Perfil arrojado, crescimento patrimonial, horizonte longo (5+ anos).",
        "quando_nao_indicar": "Perfil conservador ou moderado. Não adequado para recursos de curto prazo.",
        "alertas": "Alta volatilidade. Pode ter drawdowns significativos. Tributação ações (15% IR).",
    },
    "Dividendos": {
        "quando_indicar": "Perfil moderado a arrojado, renda passiva, longo prazo.",
        "quando_nao_indicar": "Perfil conservador. Curto prazo.",
        "alertas": "Concentração setorial frequente (utilities, bancos). Tributação ações.",
    },
    "Small Caps": {
        "quando_indicar": "Perfil arrojado, potencial de valorização superior, longo prazo.",
        "quando_nao_indicar": "Perfil conservador ou moderado. Patrimônio pequeno.",
        "alertas": "Volatilidade muito alta. Baixa liquidez dos ativos subjacentes.",
    },
    # ── FUNDOS IMOBILIÁRIOS ───────────────────────────────────────────────
    "FII": {
        "quando_indicar": "Renda passiva, diversificação imobiliária, perfil moderado a arrojado.",
        "quando_nao_indicar": "Perfil conservador. Curto prazo.",
        "alertas": "Rendimentos isentos de IR para PF. Tributação ações (20% IR no ganho de capital).",
    },
    # ── INTERNACIONAL ─────────────────────────────────────────────────────
    "Internacional": {
        "quando_indicar": "Diversificação cambial, exposição a mercados globais, perfil arrojado.",
        "quando_nao_indicar": "Perfil conservador. Sem tolerância a variação cambial.",
        "alertas": "Risco cambial adicional. Impacto do dólar/euro no retorno.",
    },
}

DIRETRIZES_KEYS = set(DIRETRIZES.keys())

# ─── PROMPTS ─────────────────────────────────────────────────────────────────

MERGE_SYSTEM = """Você é um analista especializado em fundos de investimento. \
Recebe múltiplas versões dos campos qualitativos de um mesmo fundo e deve criar uma versão única \
sintetizada. Responda SOMENTE com JSON válido."""

MERGE_PROMPT = """O fundo "{nome}" (CNPJ: {cnpj}) tem {n} versões qualitativas de diferentes plataformas.
Sintetize-as em uma única versão, priorizando a mais completa e específica.

Versões:
{versoes}

Retorne um JSON com exatamente estas chaves:
{campos}"""

VALIDATE_SYSTEM = """Você é um analista de qualidade de dados de fundos de investimento. \
Avalie se as informações qualitativas do fundo estão alinhadas com as diretrizes da subcategoria. \
Responda SOMENTE com JSON válido."""

VALIDATE_PROMPT = """Avalie o fundo "{nome}" (subcategoria: {subcategoria}).

Diretrizes esperadas para {subcategoria}:
- quando_indicar: {dir_quando_indicar}
- quando_nao_indicar: {dir_quando_nao_indicar}
- alertas: {dir_alertas}

Informações do fundo:
- quando_indicar: {quando_indicar}
- quando_nao_indicar: {quando_nao_indicar}
- alertas: {alertas}
- descricao_simples: {descricao_simples}
- descricao_tecnica: {descricao_tecnica}

Classifique como:
- OK: conteúdo alinhado com as diretrizes
- ATENÇÃO: pequenas lacunas ou imprecisões
- REVISAR: contradição direta com o perfil da subcategoria, subcategoria claramente errada, ou ausência de alerta crítico

IMPORTANT: Só marque REVISAR para contradições diretas. Diferenças de ênfase ou estilo = OK ou ATENÇÃO.

Retorne JSON: {{"validacao_status": "OK|ATENÇÃO|REVISAR", "validacao_obs": "observação curta ou vazio se OK"}}"""

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _parse_json(text: str) -> dict:
    text = text.strip()
    if "```" in text:
        for part in text.split("```"):
            part = part.strip()
            if part.startswith("json"):
                part = part[4:]
            if part.strip().startswith("{"):
                text = part.strip()
                break
    return json.loads(text)


def carregar_ck(path: str) -> dict:
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def salvar_ck(path: str, ck: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(ck, f, ensure_ascii=False, indent=2)


# ─── ETAPA 1: MESCLAGEM DE DUPLICATAS ────────────────────────────────────────

def mesclar_duplicatas(df: pd.DataFrame, client: anthropic.Anthropic) -> pd.DataFrame:
    ck = carregar_ck(CHECKPOINT_MESCLA)

    # Encontra CNPJs duplicados
    dup_cnpjs = df[df["CNPJ"].duplicated(keep=False)]["CNPJ"].unique()
    nao_dup   = df[~df["CNPJ"].isin(dup_cnpjs)].copy()
    fundos_dup = df[df["CNPJ"].isin(dup_cnpjs)].copy()

    print(f"\n🔀 Duplicatas: {len(dup_cnpjs)} CNPJs | {len(nao_dup)} fundos únicos")

    mesclados = []
    for cnpj in dup_cnpjs:
        if cnpj in ck:
            mesclados.append(ck[cnpj])
            continue

        grupo = fundos_dup[fundos_dup["CNPJ"] == cnpj]
        nome  = grupo["Nome"].iloc[0]
        print(f"  Mesclando: {nome[:50]} ({len(grupo)} versões)")

        # Escolhe linha base pela ordem de prioridade de plataforma
        base = grupo.iloc[0].to_dict()  # fallback

        versoes_str = ""
        for i, (_, row) in enumerate(grupo.iterrows()):
            versoes_str += f"\nVersão {i+1}:\n"
            for col in QUALITATIVE_COLS:
                val = row.get(col.replace("_", " ").title(), "")
                versoes_str += f"  {col}: {val}\n"

        campos_json = str({c: "string" for c in QUALITATIVE_COLS})

        try:
            resp = client.messages.create(
                model=MODEL_SONNET,
                max_tokens=2000,
                system=MERGE_SYSTEM,
                messages=[{"role": "user", "content": MERGE_PROMPT.format(
                    nome=nome, cnpj=cnpj, n=len(grupo),
                    versoes=versoes_str, campos=campos_json
                )}]
            )
            campos_mesclados = _parse_json(resp.content[0].text)
            base.update({
                col.replace("_", " ").title(): campos_mesclados.get(col, "")
                for col in QUALITATIVE_COLS
            })
        except Exception as e:
            print(f"    ⚠️  Erro mesclagem: {e} — usando primeira versão")

        ck[cnpj] = base
        salvar_ck(CHECKPOINT_MESCLA, ck)
        mesclados.append(base)
        time.sleep(0.5)

    df_mesclados = pd.DataFrame(mesclados) if mesclados else pd.DataFrame(columns=df.columns)
    return pd.concat([nao_dup, df_mesclados], ignore_index=True)


# ─── ETAPA 2: VALIDAÇÃO ───────────────────────────────────────────────────────

def validar_fundos(df: pd.DataFrame, client: anthropic.Anthropic) -> pd.DataFrame:
    ck = carregar_ck(CHECKPOINT_VALIDACAO)

    df["validacao_status"] = ""
    df["validacao_obs"]    = ""

    total = len(df)
    for idx, row in df.iterrows():
        cnpj = str(row.get("CNPJ", ""))
        if cnpj in ck:
            df.at[idx, "validacao_status"] = ck[cnpj].get("validacao_status", "")
            df.at[idx, "validacao_obs"]    = ck[cnpj].get("validacao_obs", "")
            continue

        nome = row.get("Nome", "")
        subcat = row.get("Subcategoria", "")
        print(f"  [{idx+1}/{total}] {str(nome)[:50]}")

        # Sem diretriz disponível
        if subcat not in DIRETRIZES_KEYS:
            df.at[idx, "validacao_status"] = "SEM INFO"
            df.at[idx, "validacao_obs"]    = f"Subcategoria '{subcat}' sem diretriz definida"
            ck[cnpj] = {"validacao_status": "SEM INFO", "validacao_obs": df.at[idx, "validacao_obs"]}
            salvar_ck(CHECKPOINT_VALIDACAO, ck)
            continue

        # Sem conteúdo qualitativo
        has_qual = any(row.get(col.replace("_", " ").title()) for col in QUALITATIVE_COLS)
        if not has_qual:
            df.at[idx, "validacao_status"] = "SEM INFO"
            df.at[idx, "validacao_obs"]    = "Sem conteúdo qualitativo preenchido"
            ck[cnpj] = {"validacao_status": "SEM INFO", "validacao_obs": df.at[idx, "validacao_obs"]}
            salvar_ck(CHECKPOINT_VALIDACAO, ck)
            continue

        diretriz = DIRETRIZES[subcat]
        try:
            resp = client.messages.create(
                model=MODEL_HAIKU,
                max_tokens=300,
                system=VALIDATE_SYSTEM,
                messages=[{"role": "user", "content": VALIDATE_PROMPT.format(
                    nome=nome, subcategoria=subcat,
                    dir_quando_indicar=diretriz.get("quando_indicar", ""),
                    dir_quando_nao_indicar=diretriz.get("quando_nao_indicar", ""),
                    dir_alertas=diretriz.get("alertas", ""),
                    quando_indicar=row.get("Quando Indicar", ""),
                    quando_nao_indicar=row.get("Quando não indicar", ""),
                    alertas=row.get("Alertas", ""),
                    descricao_simples=row.get("Descrição Simples", ""),
                    descricao_tecnica=row.get("Descrição Técnica", ""),
                )}]
            )
            resultado = _parse_json(resp.content[0].text)
            status = resultado.get("validacao_status", "OK")
            obs    = resultado.get("validacao_obs", "")
            df.at[idx, "validacao_status"] = status
            df.at[idx, "validacao_obs"]    = obs
            ck[cnpj] = {"validacao_status": status, "validacao_obs": obs}
            salvar_ck(CHECKPOINT_VALIDACAO, ck)
            print(f"    → {status}")
        except Exception as e:
            print(f"    ❌ Erro: {e}")
            df.at[idx, "validacao_status"] = "ERRO"
            df.at[idx, "validacao_obs"]    = str(e)

        time.sleep(0.2)

    return df


# ─── SALVAR EXCEL ────────────────────────────────────────────────────────────

def salvar_excel_final(df: pd.DataFrame, df_original: pd.DataFrame) -> None:
    wb = openpyxl.Workbook()

    # Aba 1: Fundos Revisados
    ws1 = wb.active
    ws1.title = "Fundos Revisados"
    hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")

    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws1.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        status = row.get("validacao_status", "")
        fill = None
        if status == "REVISAR":
            fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        elif status == "ATENÇÃO":
            fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        for ci, col in enumerate(cols, 1):
            cell = ws1.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    ws1.freeze_panes = "A2"

    # Aba 2: Disponibilidade
    ws2 = wb.create_sheet("Disponibilidade")
    cabecalho_disp = ["CNPJ", "Nome"] + PLATAFORMAS
    for ci, col in enumerate(cabecalho_disp, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, (_, row) in enumerate(df_original.iterrows(), 2):
        cnpj = row.get("CNPJ", "")
        nome = row.get("Nome", "")
        ws2.cell(row=ri, column=1, value=cnpj)
        ws2.cell(row=ri, column=2, value=nome)
        for ci, plat in enumerate(PLATAFORMAS, 3):
            # Marca ✓ se o fundo está disponível na plataforma
            disponivel = str(row.get(plat, "")).strip().upper() in ["X", "✓", "SIM", "1"]
            ws2.cell(row=ri, column=ci, value="✓" if disponivel else "")

    wb.save(EXCEL_OUTPUT)
    print(f"\n✅ Excel salvo: {EXCEL_OUTPUT}")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=EXCEL_INPUT)
    args = parser.parse_args()

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    print(f"📂 Carregando {args.input}...")
    df_original = pd.read_excel(args.input)
    df_original = df_original.fillna("")
    print(f"   {len(df_original)} linhas carregadas")

    # Etapa 1: Mesclar duplicatas
    df = mesclar_duplicatas(df_original, client)
    print(f"\n✅ Após mesclagem: {len(df)} fundos únicos")

    # Etapa 2: Validar
    print(f"\n🔍 Validando {len(df)} fundos...")
    df = validar_fundos(df, client)

    # Resumo
    status_counts = df["validacao_status"].value_counts()
    print(f"\n📊 Resultado da validação:")
    for status, count in status_counts.items():
        print(f"   {status}: {count}")

    # Salvar
    salvar_excel_final(df, df_original)


if __name__ == "__main__":
    main()
