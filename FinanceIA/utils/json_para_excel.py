"""
json_para_excel.py
Converte qualquer arquivo JSON para Excel formatado.

Uso:
    python json_para_excel.py fundos_xp_raw.json
    python json_para_excel.py fundos_btg_raw.json --output btg_visualizacao.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def main():
    parser = argparse.ArgumentParser(description="Converte JSON para Excel formatado")
    parser.add_argument("input", help="Arquivo JSON de entrada")
    parser.add_argument("--output", help="Arquivo Excel de saída (opcional)")
    parser.add_argument("--sheet", default="Dados", help="Nome da aba (default: Dados)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ Arquivo não encontrado: {args.input}")
        sys.exit(1)

    output_path = Path(args.output) if args.output else input_path.with_suffix(".xlsx")

    print(f"📂 Carregando {args.input}...")
    with open(input_path, encoding="utf-8") as f:
        dados = json.load(f)

    # Suporta lista de objetos ou objeto raiz
    if isinstance(dados, dict):
        # Tenta encontrar a lista dentro do dict
        for key, val in dados.items():
            if isinstance(val, list) and len(val) > 0:
                dados = val
                print(f"   Usando campo '{key}': {len(dados)} registros")
                break
        else:
            dados = [dados]

    if not isinstance(dados, list):
        print("❌ JSON deve ser uma lista de objetos ou um objeto com lista interna")
        sys.exit(1)

    print(f"   {len(dados)} registros encontrados")

    df = pd.json_normalize(dados)
    print(f"   {len(df.columns)} colunas extraídas")

    # Salva com formatação
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.sheet

    hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")

    # Cabeçalhos
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(len(str(col)), 15)

    ws.row_dimensions[1].height = 30

    # Dados
    body_font = Font(name="Arial", size=10)
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val if not pd.isna(val) else "")
            cell.font = body_font
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\n✅ Excel salvo: {output_path}")
    print(f"   {len(df)} linhas × {len(df.columns)} colunas")


if __name__ == "__main__":
    main()
