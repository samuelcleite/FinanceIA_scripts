"""
cvm_enriquecimento.py
Enriquecimento de dados de mercado via CVM (dados.cvm.gov.br)

Calcula:
- Rentabilidade 12m, 24m, 36m e desde o início (variação de cota)
- Histórico mensal de PL (último dia útil de cada mês, 36 meses)

Input:  consolidado_fundos.xlsx  (coluna "CNPJ")
Output: dados_mercado_cvm.xlsx   (aba "Dados de Mercado")

Uso:
    python cvm_enriquecimento.py
    python cvm_enriquecimento.py --input meu_arquivo.xlsx --output saida.xlsx
"""

import argparse
import io
import os
import zipfile
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import requests

# ─── CONFIGURAÇÃO ────────────────────────────────────────────────────────────

ARQUIVO_CNPJS  = "consolidado_fundos.xlsx"
COLUNA_CNPJ    = "CNPJ"
ARQUIVO_SAIDA  = "dados_mercado_cvm.xlsx"
CACHE_DIR      = Path("cache_cvm")
MESES_HISTORICO = 36   # quantos meses de PL histórico incluir

# URL base dos informes diários da CVM
CVM_BASE_URL = "https://dados.cvm.gov.br/dados/FI/DOC/INF_DIARIO/DADOS/"

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def normalizar_cnpj(cnpj: str) -> str:
    """Remove formatação do CNPJ: '00.601.692/0001-23' → '00601692000123'"""
    if not cnpj:
        return ""
    return re.sub(r'[^0-9]', '', str(cnpj))


def formatar_cnpj(cnpj: str) -> str:
    """Adiciona formatação ao CNPJ: '00601692000123' → '00.601.692/0001-23'"""
    c = normalizar_cnpj(cnpj)
    if len(c) != 14:
        return cnpj
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"


import re


def baixar_informe_mensal(ano: int, mes: int) -> pd.DataFrame | None:
    """Baixa e cacheia o informe diário mensal da CVM."""
    CACHE_DIR.mkdir(exist_ok=True)
    nome_arquivo = f"inf_diario_fi_{ano}{mes:02d}.csv"
    cache_path   = CACHE_DIR / nome_arquivo

    if cache_path.exists():
        return pd.read_csv(cache_path, sep=";", encoding="latin-1", dtype=str)

    # Tenta URL com ZIP
    url_zip = f"{CVM_BASE_URL}inf_diario_fi_{ano}{mes:02d}.zip"
    try:
        resp = requests.get(url_zip, timeout=60)
        if resp.status_code == 200:
            with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
                csv_name = [n for n in z.namelist() if n.endswith(".csv")][0]
                with z.open(csv_name) as f:
                    df = pd.read_csv(f, sep=";", encoding="latin-1", dtype=str)
            df.to_csv(cache_path, index=False, encoding="utf-8")
            print(f"    ✅ {ano}/{mes:02d}: {len(df):,} registros")
            return df
    except Exception as e:
        print(f"    ⚠️  Falha {ano}/{mes:02d}: {e}")
    return None


def calcular_rentabilidade(df_fundo: pd.DataFrame, meses: int) -> float | None:
    """
    Calcula rentabilidade por variação de cota entre dois pontos.
    df_fundo deve ter colunas: DT_COMPTC, VL_QUOTA (ordenado por data).
    """
    if df_fundo.empty:
        return None

    df_sorted = df_fundo.sort_values("DT_COMPTC")
    cota_final = df_sorted["VL_QUOTA"].iloc[-1]

    # Busca cota de N meses atrás
    data_final = pd.to_datetime(df_sorted["DT_COMPTC"].iloc[-1])
    data_ref   = data_final - pd.DateOffset(months=meses)
    df_passado = df_sorted[pd.to_datetime(df_sorted["DT_COMPTC"]) <= data_ref]

    if df_passado.empty:
        return None

    cota_inicial = df_passado["VL_QUOTA"].iloc[-1]
    try:
        cota_final_f   = float(str(cota_final).replace(",", "."))
        cota_inicial_f = float(str(cota_inicial).replace(",", "."))
        if cota_inicial_f == 0:
            return None
        return round((cota_final_f / cota_inicial_f - 1) * 100, 4)
    except (ValueError, TypeError):
        return None


def calcular_rentabilidade_inicio(df_fundo: pd.DataFrame) -> float | None:
    if len(df_fundo) < 2:
        return None
    df_sorted = df_fundo.sort_values("DT_COMPTC")
    try:
        cota_ini = float(str(df_sorted["VL_QUOTA"].iloc[0]).replace(",", "."))
        cota_fim = float(str(df_sorted["VL_QUOTA"].iloc[-1]).replace(",", "."))
        if cota_ini == 0:
            return None
        return round((cota_fim / cota_ini - 1) * 100, 4)
    except (ValueError, TypeError):
        return None


def pl_ultimo_dia_mes(df_fundo: pd.DataFrame, ano: int, mes: int) -> float | None:
    """Retorna o PL do último dia útil do mês para o fundo."""
    df_mes = df_fundo[
        (pd.to_datetime(df_fundo["DT_COMPTC"]).dt.year == ano) &
        (pd.to_datetime(df_fundo["DT_COMPTC"]).dt.month == mes)
    ]
    if df_mes.empty:
        return None
    df_sorted = df_mes.sort_values("DT_COMPTC")
    try:
        return float(str(df_sorted["VL_PATRIM_LIQ"].iloc[-1]).replace(",", "."))
    except (ValueError, TypeError):
        return None


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  default=ARQUIVO_CNPJS)
    parser.add_argument("--output", default=ARQUIVO_SAIDA)
    args = parser.parse_args()

    # 1. Carrega CNPJs
    print(f"📂 Carregando CNPJs de: {args.input}")
    df_fundos = pd.read_excel(args.input)
    cnpjs_raw = df_fundos[COLUNA_CNPJ].dropna().unique().tolist()
    cnpjs_norm = {normalizar_cnpj(str(c)): str(c) for c in cnpjs_raw}
    print(f"   {len(cnpjs_norm)} CNPJs únicos carregados")

    # 2. Define período (meses mais recentes)
    hoje = date.today()
    meses = []
    for i in range(MESES_HISTORICO):
        if hoje.month - i > 0:
            ano_ref = hoje.year
            mes_ref = hoje.month - i
        else:
            sobra = i - hoje.month
            ano_ref = hoje.year - 1 - (sobra // 12)
            mes_ref = 12 - (sobra % 12)
        meses.append((ano_ref, mes_ref))
    meses = sorted(meses)

    # 3. Baixa e consolida dados da CVM
    print(f"\n📥 Baixando {len(meses)} meses de dados da CVM...")
    df_total = pd.DataFrame()
    for (ano, mes) in meses:
        df_mes = baixar_informe_mensal(ano, mes)
        if df_mes is not None:
            # Filtra apenas CNPJs de interesse
            df_mes["CNPJ_NORM"] = df_mes["CNPJ_FUNDO"].apply(lambda x: normalizar_cnpj(str(x)))
            df_filtrado = df_mes[df_mes["CNPJ_NORM"].isin(cnpjs_norm.keys())]
            df_total = pd.concat([df_total, df_filtrado], ignore_index=True)

    if df_total.empty:
        print("❌ Nenhum dado encontrado. Verifique os CNPJs.")
        return

    print(f"\n✅ Total de registros carregados: {len(df_total):,}")

    # 4. Calcula métricas por fundo
    print(f"\n⚙️  Calculando métricas por fundo...")
    meses_labels = [f"{ano}/{mes:02d}" for (ano, mes) in meses]
    resultados = []

    for cnpj_norm, cnpj_original in cnpjs_norm.items():
        df_fundo = df_total[df_total["CNPJ_NORM"] == cnpj_norm].copy()

        row = {
            "CNPJ": formatar_cnpj(cnpj_norm),
            "Data_Ref": None,
            "Rent_12m_%":    None,
            "Rent_24m_%":    None,
            "Rent_36m_%":    None,
            "Rent_Inicio_%": None,
            "PL_Atual":      None,
            "Cotistas":      None,
            "Capt_30d":      None,
            "Resg_30d":      None,
        }

        if not df_fundo.empty:
            df_fundo_sorted = df_fundo.sort_values("DT_COMPTC")
            row["Data_Ref"]     = df_fundo_sorted["DT_COMPTC"].iloc[-1]
            row["Rent_12m_%"]   = calcular_rentabilidade(df_fundo, 12)
            row["Rent_24m_%"]   = calcular_rentabilidade(df_fundo, 24)
            row["Rent_36m_%"]   = calcular_rentabilidade(df_fundo, 36)
            row["Rent_Inicio_%"] = calcular_rentabilidade_inicio(df_fundo)

            try:
                row["PL_Atual"]  = float(str(df_fundo_sorted["VL_PATRIM_LIQ"].iloc[-1]).replace(",", "."))
                row["Cotistas"]  = int(df_fundo_sorted["NR_COTST"].iloc[-1])
            except (ValueError, TypeError):
                pass

            # Captação e resgate últimos 30 dias
            data_max = pd.to_datetime(df_fundo_sorted["DT_COMPTC"].iloc[-1])
            df_30d   = df_fundo_sorted[pd.to_datetime(df_fundo_sorted["DT_COMPTC"]) >= (data_max - pd.Timedelta(days=30))]
            try:
                row["Capt_30d"] = df_30d["CAPTC_DIA"].astype(str).str.replace(",", ".").astype(float).sum()
                row["Resg_30d"] = df_30d["RESG_DIA"].astype(str).str.replace(",", ".").astype(float).sum()
            except Exception:
                pass

            # PL mensal histórico
            for (ano, mes) in meses:
                label = f"PL_{ano}/{mes:02d}"
                row[label] = pl_ultimo_dia_mes(df_fundo, ano, mes)

        resultados.append(row)

    df_resultado = pd.DataFrame(resultados)
    com_dados  = df_resultado["Data_Ref"].notna().sum()
    sem_dados  = df_resultado["Data_Ref"].isna().sum()

    # 5. Exporta Excel
    print(f"\n📊 Exportando para: {args.output}")
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        df_resultado.to_excel(writer, sheet_name="Dados de Mercado", index=False)
        ws = writer.sheets["Dados de Mercado"]
        ws.freeze_panes = "A2"

        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter

        hdr_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)

        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 18

        ws.row_dimensions[1].height = 30
        ws.auto_filter.ref = ws.dimensions

    print(f"\n{'=' * 60}")
    print(f"  Arquivo: {args.output}")
    print(f"  Total de fundos: {len(df_resultado)}")
    print(f"  Com dados CVM:   {com_dados}")
    print(f"  Sem dados CVM:   {sem_dados}")
    print(f"  Período PL:      {meses_labels[0]} a {meses_labels[-1]}")
    print(f"{'=' * 60}")

    if sem_dados > 0:
        print(f"\n⚠️  Fundos sem dados (primeiros 10):")
        for _, r in df_resultado[df_resultado["Data_Ref"].isna()].head(10).iterrows():
            print(f"    - {r['CNPJ']}")


if __name__ == "__main__":
    main()
