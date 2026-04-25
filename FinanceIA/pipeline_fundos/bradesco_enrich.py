"""
bradesco_enrich.py
Enriquecimento de CNPJ, Liquidez e Taxa de Performance para fundos do Bradesco.
Baixa as lâminas PDF do Bradesco e extrai via Claude Sonnet.

URL padrão das lâminas: https://wspf.bradesco.com.br/wsFundosInvestimentos/File/{sistemaOrigem}_{codigoProduto}_LAMINA.pdf

Uso:
    python bradesco_enrich.py
"""

import base64
import json
import re
import time
from pathlib import Path

import anthropic
import openpyxl
import requests

# ─── CONFIGURAÇÃO ────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = "COLE_SUA_CHAVE_AQUI"   # sk-ant-api03-...
MODEL             = "claude-sonnet-4-20250514"

INPUT_XLSX  = "fundos_bradesco.xlsx"   # planilha já preenchida pelo pipeline principal
INPUT_JSON  = "fundos_bradesco_raw.json"
OUTPUT_XLSX = "fundos_bradesco_enriquecido.xlsx"
PDF_DIR     = Path("pdfs_bradesco")
CHECKPOINT  = "progress_bradesco_enrich.json"

DELAY_S = 1.5

CAMPOS_EXTRAIR = ["CNPJ", "Liquidez", "Taxa de Performance"]

SYSTEM_PROMPT = """Você é um analista especializado em fundos de investimento brasileiros. \
Extraia as informações solicitadas exatamente como estão no documento. \
Retorne apenas um objeto JSON válido, sem texto adicional, sem blocos markdown."""

EXTRACT_PROMPT = """Analise esta lâmina de fundo de investimento e retorne um JSON com exatamente estas 3 chaves:

{{
  "CNPJ": "CNPJ do fundo no formato 00.000.000/0000-00. Se não encontrar, retorne string vazia.",
  "Liquidez": "Prazo de liquidação/resgate, ex: D+1 cotização / D+1 liquidação, D+30 dias, etc. Se não encontrar, string vazia.",
  "Taxa de Performance": "Taxa de performance exatamente como consta no documento, ex: 20% sobre o que exceder o CDI. Se não houver taxa de performance, retorne Não há."
}}

Extraia apenas o que está no documento. Não invente dados."""

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _parse_json(text: str) -> dict:
    text = text.strip()
    if "```" in text:
        for part in text.split("```"):
            part = part.strip()
            if part.startswith("json"):
                part = part[4:]
            if part.strip().startswith("{"):
                text = part.strip()
                break
    return json.loads(text)


def montar_url_lamina(fundo: dict) -> str:
    sistema = fundo.get("sistemaOrigem", "")
    codigo  = fundo.get("codigoProduto", "")
    return f"https://wspf.bradesco.com.br/wsFundosInvestimentos/File/{sistema}_{codigo}_LAMINA.pdf"


def baixar_pdf(url: str, nome: str) -> bytes | None:
    PDF_DIR.mkdir(exist_ok=True)
    nome_limpo = "".join(c for c in nome if c.isalnum() or c in " _-")[:50]
    cache = PDF_DIR / f"{nome_limpo}.pdf"

    if cache.exists() and cache.stat().st_size > 1000:
        return cache.read_bytes()

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/pdf,*/*",
        "Referer": "https://fundos.bradesconetempresa.b.com.br/",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 200 and len(resp.content) > 1000:
            cache.write_bytes(resp.content)
            return resp.content
        print(f"    HTTP {resp.status_code}")
    except Exception as e:
        print(f"    Erro: {e}")
    return None


def carregar_checkpoint() -> dict:
    try:
        with open(CHECKPOINT, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def salvar_checkpoint(ck: dict) -> None:
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump(ck, f, ensure_ascii=False, indent=2)


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # Carrega JSON para ter sistemaOrigem e codigoProduto
    with open(INPUT_JSON, encoding="utf-8") as f:
        fundos_raw = json.load(f)
    nome_para_fundo = {f["nomeFundo"]: f for f in fundos_raw}

    # Carrega planilha existente
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb.active

    # Mapeia cabeçalhos
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

    # Garante que as colunas de destino existem
    for campo in CAMPOS_EXTRAIR:
        if campo not in headers:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx, value=campo)
            headers[campo] = col_idx

    ck = carregar_checkpoint()
    total = ws.max_row - 1
    ok_count = 0
    erro_count = 0

    for row_idx in range(2, ws.max_row + 1):
        nome_cell = ws.cell(row=row_idx, column=headers.get("Nome", 1))
        nome = nome_cell.value or ""

        if not nome:
            continue

        # Verifica se já foi processado
        if ck.get(nome, {}).get("_status") == "ok":
            ok_count += 1
            continue

        print(f"[{row_idx-1}/{total}] {str(nome)[:60]}")

        fundo_raw = nome_para_fundo.get(nome)
        if not fundo_raw:
            print(f"  ⚠️  Não encontrado no JSON")
            erro_count += 1
            continue

        url = montar_url_lamina(fundo_raw)
        pdf_bytes = baixar_pdf(url, nome)

        if not pdf_bytes:
            print(f"  ⚠️  Sem PDF disponível")
            ck[nome] = {"_status": "sem_pdf"}
            salvar_checkpoint(ck)
            continue

        try:
            pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
            resp = client.messages.create(
                model=MODEL, max_tokens=500, system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": [
                    {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                    {"type": "text", "text": EXTRACT_PROMPT},
                ]}]
            )
            campos = _parse_json(resp.content[0].text)

            for campo, valor in campos.items():
                if campo in headers and valor:
                    ws.cell(row=row_idx, column=headers[campo], value=valor)

            ck[nome] = {"_status": "ok", **campos}
            salvar_checkpoint(ck)
            wb.save(OUTPUT_XLSX)
            ok_count += 1
            print(f"  ✅ CNPJ: {campos.get('CNPJ', '-')} | Liq: {campos.get('Liquidez', '-')}")

        except Exception as e:
            print(f"  ❌ ERRO: {e}")
            ck[nome] = {"_status": "erro", "_erro": str(e)}
            salvar_checkpoint(ck)
            erro_count += 1

        time.sleep(DELAY_S)

    wb.save(OUTPUT_XLSX)
    print(f"\n📊 OK: {ok_count} | Erros: {erro_count}")
    print(f"💾 Salvo: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
