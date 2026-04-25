"""
itau_fund_extractor.py
Pipeline FinanceIA — Itaú Íon
Lê fundos_itau_raw.json → baixa PDFs via campo lamina_produto → gera campos analíticos via Claude API
Output: fundos_itau_preenchidos.xlsx

Uso:
    python itau_fund_extractor.py
    python itau_fund_extractor.py --test 5
    python itau_fund_extractor.py --reset
"""

import argparse
import base64
import json
import logging
import re
import time
from pathlib import Path

import anthropic
import openpyxl
import pdfplumber
import requests
from openpyxl.styles import Alignment, Font, PatternFill

# ─── CONFIGURAÇÃO ────────────────────────────────────────────────────────────

INPUT_JSON    = "fundos_itau_raw.json"
OUTPUT_FILE   = "fundos_itau_preenchidos.xlsx"
PDF_LOCAL_DIR = Path("pdfs_itau")
PROGRESS_FILE = "progress_itau.json"

ANTHROPIC_API_KEY = "COLE_SUA_CHAVE_AQUI"   # sk-ant-api03-...
MODEL             = "claude-sonnet-4-20250514"
DELAY_S           = 1.5

CNPJ_REGEX = re.compile(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}')

# Mapeamento de classe Itaú → Categoria do Modelo
CLASSE_ITAU_TO_CATEGORIA = {
    "RENDA FIXA":        "Renda Fixa",
    "MULTIMERCADO":      "Multimercado",
    "AÇÕES":             "Renda Variável",
    "ACOES":             "Renda Variável",
    "CAMBIAL":           "Cambial",
    "PREVIDENCIA":       "Previdência",
    "FII":               "Fundos Imobiliários",
    "INTERNACIONAL":     "Internacional",
}

COLUNAS_MODELO = [
    "Nome", "CNPJ", "Categoria", "Subcategoria", "Gestor",
    "Indexador", "Liquidez", "Tributação", "Come-Cotas",
    "Taxa de Adm", "Taxa de Performance", "Benchmark",
    "Descrição Tributação", "Público Alvo", "Horizonte Mínimo (anos)",
    "Quando Indicar", "Quando não indicar", "Vantagens",
    "Desvantagens", "Alertas", "Descrição Simples", "Descrição Técnica",
    "_status", "_erro",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler("extrator_itau.log", encoding="utf-8"), logging.StreamHandler()]
)
log = logging.getLogger(__name__)

# ─── PROMPTS ─────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Você é um especialista em fundos de investimento brasileiros com certificação CFP, \
atuando como analista editorial da 3A Riva Investimentos.

Analise lâminas e regulamentos de fundos e produza conteúdo objetivo, específico e acionável para \
assessores de investimento. Evite afirmações vagas. Nunca mencione rentabilidade passada como garantia.

Responda SOMENTE com um objeto JSON válido, sem texto adicional, sem blocos de código markdown."""

ANALYSIS_PROMPT = """Analise a lâmina/regulamento deste fundo e retorne um JSON com exatamente estas 9 chaves.

Contexto estruturado (Itaú):
- Nome: {nome}
- Categoria: {categoria}
- Subcategoria: {subcategoria}
- Benchmark: {benchmark}
- Taxa de Adm: {taxa_adm}
- Liquidez: {liquidez}
- Aplicação Mínima: R$ {aplic_minima}
- Tributação: {tributacao}
- Come-Cotas: {come_cotas}

Escala de horizonte: Curto (até 1 ano) | Médio (1 a 3 anos) | Longo (3+ anos)
Objetivos: Aposentadoria | Reserva de emergência | Compra de imóvel | Independência financeira | Crescimento patrimonial

JSON esperado:
{{
  "Gestor": "Nome completo da gestora extraído do documento. Máx 40 chars.",
  "Horizonte Mínimo (anos)": "Curto (até 1 ano)" | "Médio (1 a 3 anos)" | "Longo (3+ anos)",
  "Quando Indicar": "Para qual perfil e objetivo. Cruze risco × horizonte × objetivo. Máx 80 palavras.",
  "Quando não indicar": "Perfis/situações inadequadas. Cite volatilidade se disponível no PDF. Máx 60 palavras.",
  "Vantagens": "3 vantagens concretas vs média da categoria. Com números quando possível. Máx 80 palavras.",
  "Desvantagens": "2-3 desvantagens reais. Máx 60 palavras.",
  "Alertas": "Prioridade: (1) rentabilidade abaixo benchmark — cite números; (2) PL baixo; (3) liquidez longa. Máx 80 palavras.",
  "Descrição Simples": "2-3 frases para leigo. O que faz, objetivo, para quem. Máx 60 palavras.",
  "Descrição Técnica": "Estratégia, ativos, benchmarks, instrumentos, riscos. Para o assessor. Máx 120 palavras."
}}"""

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _parse_json(text: str) -> dict:
    text = text.strip()
    if "```" in text:
        for part in text.split("```"):
            part = part.strip()
            if part.startswith("json"):
                part = part[4:]
            if part.strip().startswith("{"):
                text = part.strip()
                break
    return json.loads(text)


def extrair_cnpj_do_pdf(pdf_path: Path) -> str:
    """Extrai CNPJ da primeira página do PDF via regex. Sem IA."""
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            texto = pdf.pages[0].extract_text() or ""
            matches = CNPJ_REGEX.findall(texto)
            return matches[0] if matches else ""
    except Exception as e:
        log.warning(f"  Erro ao extrair CNPJ: {e}")
        return ""


def baixar_pdf(url: str, nome: str) -> Path | None:
    PDF_LOCAL_DIR.mkdir(exist_ok=True)
    nome_limpo = "".join(c for c in nome if c.isalnum() or c in " _-")[:50]
    caminho = PDF_LOCAL_DIR / f"{nome_limpo}.pdf"

    if caminho.exists() and caminho.stat().st_size > 1000:
        return caminho

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/pdf,*/*",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        if resp.status_code == 200 and len(resp.content) > 1000:
            caminho.write_bytes(resp.content)
            log.info(f"  PDF baixado: {len(resp.content)//1024} KB")
            return caminho
        log.warning(f"  Falha PDF: HTTP {resp.status_code}")
    except Exception as e:
        log.error(f"  Erro download: {e}")
    return None


def analisar_fundo(client: anthropic.Anthropic, fundo: dict, pdf_path: Path) -> dict:
    pdf_bytes = pdf_path.read_bytes()
    pdf_b64   = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    categoria = CLASSE_ITAU_TO_CATEGORIA.get(
        fundo.get("classeAtivo", "").upper(), fundo.get("classeAtivo", "")
    )

    prompt = ANALYSIS_PROMPT.format(
        nome=fundo.get("nomeFundo", ""),
        categoria=categoria,
        subcategoria=fundo.get("subClasse", ""),
        benchmark=fundo.get("indicadorReferencia", ""),
        taxa_adm=fundo.get("taxaAdministracao", ""),
        liquidez=fundo.get("liquidez", ""),
        aplic_minima=fundo.get("valorMinimoAplicacao", ""),
        tributacao=fundo.get("tributacao", ""),
        come_cotas="Sim" if fundo.get("comeCota") else "Não",
    )

    resp = client.messages.create(
        model=MODEL,
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                {"type": "text", "text": prompt},
            ]
        }]
    )
    return _parse_json(resp.content[0].text)


# ─── CHECKPOINT ──────────────────────────────────────────────────────────────

def carregar_checkpoint() -> dict:
    try:
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def salvar_checkpoint(ck: dict) -> None:
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(ck, f, ensure_ascii=False, indent=2)


# ─── EXCEL ───────────────────────────────────────────────────────────────────

def salvar_excel(resultados: list) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fundos Itaú"

    hdr_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")

    for ci, col in enumerate(COLUNAS_MODELO, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(resultados, 2):
        status = row.get("_status", "")
        fill = None
        if status == "erro":
            fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        elif status == "sem_pdf":
            fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")

        for ci, col in enumerate(COLUNAS_MODELO, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Excel salvo: {OUTPUT_FILE}")


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--test",  type=int, default=0)
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    with open(INPUT_JSON, encoding="utf-8") as f:
        fundos = json.load(f)
    print(f"📂 {len(fundos)} fundos carregados")

    ck = {} if args.reset else carregar_checkpoint()
    processados = set(ck.keys())
    pendentes = [f for f in fundos if f.get("codigoProduto", "") not in processados]
    if args.test:
        pendentes = pendentes[:args.test]

    print(f"📋 {len(pendentes)} pendentes | {len(processados)} já processados\n")

    for i, fundo in enumerate(pendentes, 1):
        nome   = fundo.get("nomeFundo", f"fundo_{i}")
        codigo = fundo.get("codigoProduto", f"cod_{i}")
        print(f"[{i}/{len(pendentes)}] {nome[:60]}")

        resultado = {
            "Nome":     nome,
            "_status":  "ok",
            "_erro":    "",
        }

        # Campos diretos
        categoria = CLASSE_ITAU_TO_CATEGORIA.get(fundo.get("classeAtivo","").upper(), fundo.get("classeAtivo",""))
        resultado.update({
            "Categoria":    categoria,
            "Subcategoria": fundo.get("subClasse", ""),
            "Benchmark":    fundo.get("indicadorReferencia", ""),
            "Taxa de Adm":  fundo.get("taxaAdministracao", ""),
            "Liquidez":     fundo.get("liquidez", ""),
            "Tributação":   fundo.get("tributacao", ""),
            "Come-Cotas":   "Sim" if fundo.get("comeCota") else "Não",
            "Público Alvo": fundo.get("publicoAlvo", ""),
        })

        try:
            pdf_url = fundo.get("lamina_produto")
            if not pdf_url:
                resultado["_status"] = "sem_pdf"
            else:
                pdf_path = baixar_pdf(pdf_url, nome)
                if not pdf_path:
                    resultado["_status"] = "sem_pdf"
                    resultado["_erro"]   = "Falha no download"
                else:
                    # Extrai CNPJ via regex (sem IA)
                    cnpj = extrair_cnpj_do_pdf(pdf_path)
                    resultado["CNPJ"] = cnpj

                    campos_ia = analisar_fundo(client, fundo, pdf_path)
                    resultado.update(campos_ia)
        except Exception as e:
            resultado["_status"] = "erro"
            resultado["_erro"]   = str(e)
            log.error(f"  ERRO: {e}")

        ck[codigo] = resultado
        salvar_checkpoint(ck)
        time.sleep(DELAY_S)

    salvar_excel(list(ck.values()))
    ok    = sum(1 for r in ck.values() if r.get("_status") == "ok")
    erros = sum(1 for r in ck.values() if r.get("_status") == "erro")
    print(f"\n📊 Total: {len(ck)} | ✅ {ok} | ❌ {erros}")


if __name__ == "__main__":
    main()


# ─── SCRIPT AUXILIAR: Enriquecimento de CNPJ ─────────────────────────────────
# Rode este script SE o pipeline já foi executado sem extrair o CNPJ.
# Ele lê a planilha existente, encontra o PDF em cache e extrai o CNPJ por regex.

def enriquecer_cnpj():
    """Script separado para extrair CNPJs de planilha já preenchida."""
    import pandas as pd

    print("📂 Carregando planilha existente...")
    df = pd.read_excel(OUTPUT_FILE)

    with open(INPUT_JSON, encoding="utf-8") as f:
        fundos = json.load(f)
    nome_para_codigo = {f["nomeFundo"]: f["codigoProduto"] for f in fundos}

    extraidos = 0
    nao_encontrado = 0

    for idx, row in df.iterrows():
        if pd.notna(row.get("CNPJ")) and str(row.get("CNPJ", "")).strip():
            continue  # já tem CNPJ

        nome = row.get("Nome", "")
        codigo = nome_para_codigo.get(nome, "")
        if not codigo:
            nao_encontrado += 1
            continue

        # Busca PDF em cache
        nome_limpo = "".join(c for c in nome if c.isalnum() or c in " _-")[:50]
        pdf_path = PDF_LOCAL_DIR / f"{nome_limpo}.pdf"

        if not pdf_path.exists():
            nao_encontrado += 1
            continue

        cnpj = extrair_cnpj_do_pdf(pdf_path)
        if cnpj:
            df.at[idx, "CNPJ"] = cnpj
            extraidos += 1
            print(f"  ✅ {nome[:50]}: {cnpj}")
        else:
            nao_encontrado += 1
            print(f"  ⚠️  {nome[:50]}: CNPJ não encontrado no PDF")

    df.to_excel(OUTPUT_FILE, index=False)
    print(f"\n📊 CNPJs extraídos: {extraidos} | Não encontrados: {nao_encontrado}")
    print(f"💾 Salvo em: {OUTPUT_FILE}")
