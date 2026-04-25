"""
santander_fund_extractor.py + santander_analysis_pipeline.py
Pipeline FinanceIA — Santander Asset Management

ETAPA 1 — Extração (santander_fund_extractor):
    Pagina a API pública do Santander AM e salva fundos_santander_raw.json (78 fundos)

ETAPA 2 — Análise (santander_analysis_pipeline):
    Baixa PDFs das lâminas, extrai CNPJ via regex (pdfplumber), gera qualitativos via Claude Sonnet

Uso:
    python santander_fund_extractor.py                     # Extrai lista
    python santander_analysis_pipeline.py                  # Analisa
    python santander_analysis_pipeline.py --test 3         # Teste com 3 fundos
"""

# ══════════════════════════════════════════════════════════════
# PARTE 1 — EXTRACTOR
# ══════════════════════════════════════════════════════════════

import argparse
import base64
import json
import re
import time
from pathlib import Path

import anthropic
import openpyxl
import pdfplumber
import requests
from openpyxl.styles import Alignment, Font, PatternFill

# ─── CONFIG EXTRACTOR ────────────────────────────────────────────────────────

OUTPUT_JSON = "fundos_santander_raw.json"

# Cookies da sessão para a API de listagem (santanderassetmanagement.com.br)
# Como obter: F12 na página santanderassetmanagement.com.br/tools → Rede → clique em results?risk=... → Cookie
SESSION_COOKIES = """COLE_AQUI_O_COOKIE_DA_SESSAO_DO_SANTANDER_AM"""

API_URL = "https://www.santanderassetmanagement.com.br/api/v1/tools/finder/results"
API_PARAMS_BASE = {
    "risk": "0,7",
    "market": "1",
    "investment": "0,500000",
    "exclusive": "0",
    "page": 1,
    "size": 30,
}


def extrair_fundos_santander():
    """Extrai todos os fundos da API pública do Santander AM."""
    cookies = {}
    for item in SESSION_COOKIES.strip().split(";"):
        item = item.strip()
        if "=" in item:
            k, v = item.split("=", 1)
            cookies[k.strip()] = v.strip()

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
        "Referer": "https://www.santanderassetmanagement.com.br/tools",
    }

    todos_os_fundos = []
    pagina = 1

    while True:
        params = {**API_PARAMS_BASE, "page": pagina}
        resp = requests.get(API_URL, params=params, headers=headers, cookies=cookies, timeout=30)
        resp.raise_for_status()
        data = resp.json()

        fundos = data.get("results", data.get("data", []))
        if not fundos:
            break

        todos_os_fundos.extend(fundos)
        print(f"  Página {pagina}: {len(fundos)} fundos (total: {len(todos_os_fundos)})")

        total = data.get("total", 0)
        if len(todos_os_fundos) >= total:
            break
        pagina += 1
        time.sleep(0.5)

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(todos_os_fundos, f, ensure_ascii=False, indent=2)
    print(f"\n✅ {len(todos_os_fundos)} fundos salvos em {OUTPUT_JSON}")
    return todos_os_fundos


# ══════════════════════════════════════════════════════════════
# PARTE 2 — ANALYSIS PIPELINE
# ══════════════════════════════════════════════════════════════

ANTHROPIC_API_KEY = "COLE_SUA_CHAVE_AQUI"   # sk-ant-api03-...
MODEL             = "claude-sonnet-4-20250514"
DELAY_S           = 1.5

INPUT_JSON_ANALYSIS  = "fundos_santander_raw.json"
OUTPUT_EXCEL         = "fundos_santander_analisados.xlsx"
PDF_CACHE_DIR        = Path("pdfs_santander")
PROGRESS_FILE_ANALYSIS = "progress_santander.json"

CNPJ_REGEX = re.compile(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}')

# Cookies para download dos PDFs (santander.com.br — domínio diferente da API)
# Como obter: F12 em qualquer página do santander.com.br logado → Cookie
PDF_COOKIES = """COLE_AQUI_O_COOKIE_DO_SANTANDER_COM_BR"""

# Mapeamento family → campos do Modelo
FAMILY_MAP = {
    "Renda Fixa Soberano": {
        "Categoria": "Renda Fixa", "Subcategoria": "Soberano",
        "Indexador": "CDI", "Benchmark": "CDI",
        "Tributação": "Longo Prazo", "Come-Cotas": "Sim",
        "Horizonte Mínimo (anos)": "Curto (até 1 ano)",
        "Descrição Tributação": "IR regressivo: 22,5% (até 180 dias) a 15% (acima de 720 dias). Come-cotas semestral.",
    },
    "Multimercado": {
        "Categoria": "Multimercado", "Subcategoria": "Macro",
        "Indexador": "CDI", "Benchmark": "CDI",
        "Tributação": "Longo Prazo", "Come-Cotas": "Sim",
        "Horizonte Mínimo (anos)": "Médio (1 a 3 anos)",
        "Descrição Tributação": "IR regressivo: 22,5% (até 180 dias) a 15% (acima de 720 dias). Come-cotas semestral.",
    },
    "Renda Variável": {
        "Categoria": "Renda Variável", "Subcategoria": "Long Only",
        "Indexador": "IBOVESPA", "Benchmark": "IBOVESPA",
        "Tributação": "Ações", "Come-Cotas": "Não",
        "Horizonte Mínimo (anos)": "Longo (3+ anos)",
        "Descrição Tributação": "IR 15% sobre ganhos. Sem come-cotas. IOF nos primeiros 30 dias.",
    },
    "Previdência": {
        "Categoria": "Previdência", "Subcategoria": "PGBL/VGBL",
        "Indexador": "CDI", "Benchmark": "CDI",
        "Tributação": "Previdência", "Come-Cotas": "Não",
        "Horizonte Mínimo (anos)": "Longo (3+ anos)",
        "Descrição Tributação": "Tributação exclusiva na fonte: progressiva ou regressiva conforme plano.",
    },
}

COLUNAS_MODELO = [
    "Nome", "CNPJ", "Categoria", "Subcategoria", "Gestor",
    "Indexador", "Liquidez", "Tributação", "Come-Cotas",
    "Taxa de Adm", "Taxa de Performance", "Benchmark",
    "Descrição Tributação", "Público Alvo", "Horizonte Mínimo (anos)",
    "Quando Indicar", "Quando não indicar", "Vantagens",
    "Desvantagens", "Alertas", "Descrição Simples", "Descrição Técnica",
]
EXTRA_COLS = ["_ytd", "_yearOne", "_pl1y", "_nav", "_fonte_pdf", "_status", "_erro"]
ALL_COLUMNS = COLUNAS_MODELO + EXTRA_COLS

SYSTEM_PROMPT = """Você é um especialista em fundos de investimento brasileiros com certificação CFP, \
atuando como analista editorial da 3A Riva Investimentos.

Analise lâminas e regulamentos de fundos e produza conteúdo objetivo, específico e acionável para \
assessores de investimento. Evite afirmações vagas. Nunca mencione rentabilidade passada como garantia.

Responda SOMENTE com um objeto JSON válido, sem texto adicional, sem blocos de código markdown."""

ANALYSIS_PROMPT = """Analise a lâmina deste fundo e retorne um JSON com exatamente estas 8 chaves.

Contexto estruturado (Santander AM):
- Nome: {nome}
- Família: {family}
- Taxa de Adm: {taxa_adm}
- YTD: {ytd}% | Retorno 1 ano: {year_one}%
- NAV (cota atual): {nav}
- PL 1 ano: {pl1y}

Escala de horizonte: Curto (até 1 ano) | Médio (1 a 3 anos) | Longo (3+ anos)
Objetivos: Aposentadoria | Reserva de emergência | Compra de imóvel | Independência financeira | Crescimento patrimonial

JSON:
{{
  "Gestor": "Nome completo da gestora. Máx 40 chars.",
  "Quando Indicar": "Para qual perfil e objetivo. Máx 80 palavras.",
  "Quando não indicar": "Perfis/situações inadequadas. Máx 60 palavras.",
  "Vantagens": "3 vantagens concretas com números. Máx 80 palavras.",
  "Desvantagens": "2-3 desvantagens reais. Máx 60 palavras.",
  "Alertas": "Prioridade: (1) rentabilidade abaixo benchmark; (2) PL baixo; (3) liquidez longa. Máx 80 palavras.",
  "Descrição Simples": "2-3 frases para leigo. Máx 60 palavras.",
  "Descrição Técnica": "Estratégia, ativos, benchmarks, instrumentos, riscos. Máx 120 palavras."
}}"""

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _parse_json(text: str) -> dict:
    text = text.strip()
    if "```" in text:
        for part in text.split("```"):
            part = part.strip()
            if part.startswith("json"):
                part = part[4:]
            if part.strip().startswith("{"):
                text = part.strip()
                break
    return json.loads(text)


def extrair_cnpj(pdf_path: Path) -> str:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            text = pdf.pages[0].extract_text() or ""
            matches = CNPJ_REGEX.findall(text)
            return matches[0] if matches else ""
    except Exception:
        return ""


def baixar_pdf_santander(url: str, nome: str) -> Path | None:
    PDF_CACHE_DIR.mkdir(exist_ok=True)
    nome_limpo = "".join(c for c in nome if c.isalnum() or c in " _-")[:50]
    caminho = PDF_CACHE_DIR / f"{nome_limpo}.pdf"

    if caminho.exists() and caminho.stat().st_size > 1000:
        return caminho

    cookies = {}
    for item in PDF_COOKIES.strip().split(";"):
        item = item.strip()
        if "=" in item:
            k, v = item.split("=", 1)
            cookies[k.strip()] = v.strip()

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/pdf,*/*",
        "Referer": "https://www.santanderassetmanagement.com.br/",
    }
    try:
        resp = requests.get(url, headers=headers, cookies=cookies, timeout=30)
        if resp.status_code == 200 and len(resp.content) > 1000:
            caminho.write_bytes(resp.content)
            return caminho
        print(f"  ⚠️  HTTP {resp.status_code} ao baixar PDF")
    except Exception as e:
        print(f"  ⚠️  Erro download: {e}")
    return None


def carregar_checkpoint() -> dict:
    try:
        with open(PROGRESS_FILE_ANALYSIS, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def salvar_checkpoint(ck: dict) -> None:
    with open(PROGRESS_FILE_ANALYSIS, "w", encoding="utf-8") as f:
        json.dump(ck, f, ensure_ascii=False, indent=2)


def salvar_excel_santander(rows: list) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fundos Santander"

    hdr_fill = PatternFill(start_color="EC0000", end_color="EC0000", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")
    for ci, col in enumerate(ALL_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(ALL_COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_EXCEL)
    print(f"\n✅ Excel salvo: {OUTPUT_EXCEL}")


def analisar_santander(modo_teste: int = 0):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    with open(INPUT_JSON_ANALYSIS, encoding="utf-8") as f:
        fundos = json.load(f)

    if modo_teste:
        fundos = fundos[:modo_teste]

    ck = carregar_checkpoint()
    processados = set(ck.keys())
    pendentes = [f for f in fundos if f.get("id", f.get("name")) not in processados]
    total = len(pendentes)
    print(f"📋 {total} pendentes | {len(processados)} já processados")

    for i, fundo in enumerate(pendentes, 1):
        nome       = fundo.get("name", f"fundo_{i}")
        fundo_id   = fundo.get("id", nome)
        family     = fundo.get("family", "")
        pdf_url    = fundo.get("detail_uri", "")

        print(f"\n[{i}/{total}] {nome[:60]}")

        resultado = {col: "" for col in ALL_COLUMNS}
        resultado["Nome"] = nome
        resultado["Taxa de Adm"] = fundo.get("rate", "")
        resultado["_ytd"]     = fundo.get("ytd", "")
        resultado["_yearOne"] = fundo.get("yearOne", "")
        resultado["_pl1y"]    = fundo.get("pl1y", "")
        resultado["_nav"]     = fundo.get("nav", "")

        # Campos do family map
        mapa = FAMILY_MAP.get(family, {})
        for campo, valor in mapa.items():
            resultado[campo] = valor

        try:
            if pdf_url:
                pdf_path = baixar_pdf_santander(pdf_url, nome)
                if pdf_path:
                    resultado["CNPJ"] = extrair_cnpj(pdf_path)
                    pdf_bytes = pdf_path.read_bytes()
                    pdf_b64   = base64.standard_b64encode(pdf_bytes).decode("utf-8")
                    prompt = ANALYSIS_PROMPT.format(
                        nome=nome, family=family,
                        taxa_adm=resultado["Taxa de Adm"],
                        ytd=resultado["_ytd"], year_one=resultado["_yearOne"],
                        nav=resultado["_nav"], pl1y=resultado["_pl1y"],
                    )
                    resp = client.messages.create(
                        model=MODEL, max_tokens=2000, system=SYSTEM_PROMPT,
                        messages=[{"role": "user", "content": [
                            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                            {"type": "text", "text": prompt},
                        ]}]
                    )
                    campos = _parse_json(resp.content[0].text)
                    resultado.update(campos)
                    resultado["_fonte_pdf"] = "lamina"
                    resultado["_status"]    = "ok"
                    print(f"  ✅ OK")
                else:
                    resultado["_status"] = "sem_pdf"
                    resultado["_erro"]   = "Falha no download"
            else:
                resultado["_status"] = "sem_pdf"

        except Exception as e:
            resultado["_status"] = "erro"
            resultado["_erro"]   = str(e)
            print(f"  ❌ ERRO: {e}")

        ck[fundo_id] = resultado
        salvar_checkpoint(ck)
        time.sleep(DELAY_S)

    salvar_excel_santander(list(ck.values()))
    ok = sum(1 for r in ck.values() if r.get("_status") == "ok")
    print(f"\n📊 Total: {len(ck)} | ✅ {ok}")


# ─── ENTRY POINT ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--extrair", action="store_true", help="Etapa 1: extrair lista de fundos")
    parser.add_argument("--analisar", action="store_true", help="Etapa 2: analisar fundos com Claude")
    parser.add_argument("--test", type=int, default=0)
    args = parser.parse_args()

    if args.extrair:
        extrair_fundos_santander()
    elif args.analisar:
        analisar_santander(modo_teste=args.test)
    else:
        # Por padrão, roda a análise
        analisar_santander(modo_teste=args.test)
