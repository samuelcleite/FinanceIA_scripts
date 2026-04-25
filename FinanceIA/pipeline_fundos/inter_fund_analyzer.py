"""
inter_fund_analyzer.py
Gera as 7 colunas analíticas para os fundos exclusivos do Inter
usando os dados do JSON como contexto (sem PDF — Inter não disponibiliza lâminas via URL pública).

Uso:
    python inter_fund_analyzer.py
"""

import json
import time
import re
import openpyxl
from openpyxl.styles import Alignment
import anthropic

# ─── CONFIG ──────────────────────────────────────────────────────────────────

INPUT_XLSX  = "fundos_inter_exclusivos.xlsx"    # planilha com campos estruturais já preenchidos
OUTPUT_XLSX = "fundos_inter_exclusivos_analisados.xlsx"
JSON_PATH   = "fundos_inter_raw.json"

MODEL       = "claude-haiku-4-5-20251001"  # Haiku — custo eficiente para volume alto
MAX_TOKENS  = 2500
DELAY_S     = 1.0   # pausa entre chamadas

ANALYTICAL_COLS = [
    "Quando Indicar",
    "Quando não indicar",
    "Vantagens",
    "Desvantagens",
    "Alertas",
    "Descrição Simples",
    "Descrição Técnica",
]

# ─── PROMPTS ─────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Você é um especialista em fundos de investimento brasileiros com certificação CFP, \
atuando como analista editorial da 3A Riva Investimentos.

Sua função é analisar dados estruturados de fundos e produzir conteúdo objetivo, específico e acionável \
para assessores de investimento usarem no atendimento a clientes.

Escreva sempre com precisão técnica, mas sem exageros. Evite afirmações vagas como "boa gestão" ou \
"equipe experiente" sem embasamento concreto nos dados fornecidos. Nunca mencione rentabilidade passada \
como garantia de retorno futuro.

Responda SOMENTE com um objeto JSON válido, sem texto adicional, sem blocos de código markdown."""

ANALYSIS_PROMPT = """Analise os dados estruturados deste fundo e retorne um JSON com exatamente estas 7 chaves.

Dados disponíveis:
- Nome: {nome}
- Gestor: {gestor}
- Categoria: {categoria}
- Subcategoria: {subcategoria}
- Benchmark: {benchmark}
- Grau de Risco (1-7): {grau_risco}
- Tributação: {tributacao}
- Come-Cotas: {come_cotas}
- Taxa de Adm: {taxa_adm}% a.a.
- Liquidez: {liquidez}
- Aplicação Mínima: R$ {aplic_minima}
- Rentabilidade mês: {rent_mes}% | Ano: {rent_ano}% | 12m: {rent_12m}%
- PL: R$ {pl}

Escala de horizonte: Curto (até 1 ano) | Médio (1 a 3 anos) | Longo (3+ anos)
Objetivos do investidor: Aposentadoria | Reserva de emergência | Compra de imóvel | \
Independência financeira / renda passiva | Crescimento patrimonial

JSON esperado:
{{
  "Quando Indicar": "Para qual perfil e objetivo. Cruze grau de risco (1-3 conservador, 4-5 moderado, 6-7 arrojado) \
com horizonte e objetivos de vida. Cite o grau de risco se relevante. Máx 80 palavras.",

  "Quando não indicar": "Perfis/situações onde este fundo não é adequado. Cruze grau de risco com perfil do investidor. \
Para fundos de grau 6-7 especifique que não é para conservadores. Para fundos de liquidez longa, mencione que não é \
adequado para quem pode precisar do dinheiro em curto prazo. Máx 60 palavras.",

  "Vantagens": "3 vantagens concretas frente à média da categoria Inter. Destaque: taxa de adm competitiva, \
boa liquidez para a categoria, consistência de retorno (use rent_12m vs benchmarks do segmento), \
benchmark superado. Use os números disponíveis. Máx 80 palavras.",

  "Desvantagens": "2-3 desvantagens reais: taxa de adm acima da média da categoria, liquidez longa, \
come-cotas, PL pequeno para a estratégia, ausência de track record longo. Máx 60 palavras.",

  "Alertas": "Alerta (1): se rent_12m < benchmark esperado para a categoria — cite os números exatos; \
(2) PL abaixo de R$ 50M — sinal de atenção para a maioria das categorias; \
(3) liquidez longa (D+30 ou mais); (4) grau de risco 6-7 — só para perfis arrojados. \
Se nenhum alerta relevante: 'Sem alertas relevantes'. Máx 80 palavras.",

  "Descrição Simples": "2-3 frases para o cliente leigo. O que o fundo faz, seu objetivo e para quem serve. \
Sem jargão técnico. Máx 60 palavras.",

  "Descrição Técnica": "Estratégia de investimento, principais ativos, benchmark, tributação, liquidez, \
diferenciais e riscos principais. Para uso do assessor. Máx 120 palavras."
}}"""

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def _parse_json(text: str) -> dict:
    text = text.strip()
    if "```" in text:
        for part in text.split("```"):
            part = part.strip()
            if part.startswith("json"):
                part = part[4:]
            if part.strip().startswith("{"):
                text = part.strip()
                break
    return json.loads(text)


def construir_contexto(row: dict, json_data: dict) -> dict:
    nome = row.get("Nome", "")
    cnpj = row.get("CNPJ", "")
    fundo_json = json_data.get(cnpj, {})

    return {
        "nome":       nome,
        "gestor":     row.get("Gestor", fundo_json.get("manager", "")),
        "categoria":  row.get("Categoria", ""),
        "subcategoria": row.get("Subcategoria", ""),
        "benchmark":  row.get("Benchmark", fundo_json.get("benchmark", "")),
        "grau_risco": fundo_json.get("degreeRisk", row.get("_grau_risco", "")),
        "tributacao": row.get("Tributação", ""),
        "come_cotas": row.get("Come-Cotas", ""),
        "taxa_adm":   row.get("Taxa de Adm", "").replace("%", ""),
        "liquidez":   row.get("Liquidez", ""),
        "aplic_minima": fundo_json.get("minimumInitialInvestment", ""),
        "rent_mes":   fundo_json.get("rent_mes", ""),
        "rent_ano":   fundo_json.get("rent_ano", ""),
        "rent_12m":   fundo_json.get("rent_12m", ""),
        "pl":         fundo_json.get("netAssetValue", ""),
    }


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    client = anthropic.Anthropic(api_key="COLE_SUA_CHAVE_AQUI")

    # Carrega JSON do Inter para cruzar dados extras
    json_por_cnpj = {}
    try:
        with open(JSON_PATH, encoding="utf-8") as f:
            raw = json.load(f)
        for item in raw:
            cnpj = item.get("cnpj", "")
            if cnpj:
                json_por_cnpj[cnpj] = item
        print(f"✅ JSON carregado: {len(json_por_cnpj)} fundos")
    except FileNotFoundError:
        print(f"⚠️  {JSON_PATH} não encontrado — usando apenas dados do Excel")

    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb.active

    # Mapeia cabeçalhos
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    col_status = headers.get("_status_analise", None)

    # Adiciona coluna de status se não existir
    if col_status is None:
        col_status = ws.max_column + 1
        ws.cell(row=1, column=col_status, value="_status_analise")

    total = ws.max_row - 1
    processados = 0
    erros = 0

    for row_idx in range(2, ws.max_row + 1):
        status_atual = ws.cell(row=row_idx, column=col_status).value
        if status_atual == "ok":
            processados += 1
            continue

        row_data = {
            ws.cell(row=1, column=c).value: ws.cell(row=row_idx, column=c).value
            for c in range(1, ws.max_column)
        }
        nome = row_data.get("Nome", f"linha {row_idx}")
        cnpj = str(row_data.get("CNPJ", ""))
        print(f"[{row_idx-1}/{total}] {str(nome)[:60]}")

        # Verifica se campos analíticos já estão preenchidos
        ja_preenchido = all(row_data.get(c) for c in ANALYTICAL_COLS[:3])
        if ja_preenchido:
            ws.cell(row=row_idx, column=col_status, value="ok")
            processados += 1
            continue

        try:
            contexto = construir_contexto(row_data, json_por_cnpj)
            prompt   = ANALYSIS_PROMPT.format(**contexto)
            resp = client.messages.create(
                model=MODEL,
                max_tokens=MAX_TOKENS,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            )
            campos = _parse_json(resp.content[0].text)

            for col_name, valor in campos.items():
                if col_name in headers:
                    ws.cell(row=row_idx, column=headers[col_name], value=valor)
                    ws.cell(row=row_idx, column=headers[col_name]).alignment = Alignment(
                        vertical="top", wrap_text=True
                    )

            ws.cell(row=row_idx, column=col_status, value="ok")
            processados += 1
            print(f"  ✅ OK")

        except Exception as e:
            ws.cell(row=row_idx, column=col_status, value=f"erro: {e}")
            erros += 1
            print(f"  ❌ ERRO: {e}")

        wb.save(OUTPUT_XLSX)
        time.sleep(DELAY_S)

    wb.save(OUTPUT_XLSX)
    print(f"\n📊 Total: {total} | ✅ {processados} | ❌ {erros}")
    print(f"💾 Salvo em: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
