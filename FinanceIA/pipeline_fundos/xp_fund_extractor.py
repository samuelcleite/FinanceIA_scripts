"""
Pipeline FinanceIA — XP Investimentos
Preenchimento do Modelo de Produtos
Lê planilha XP raw → preenche campos diretos → gera campos analíticos via Claude API (PDF)
Output: Modelo_preenchido.xlsx pronto para revisão e importação no Base44

Uso:
    py xp_fund_extractor.py
    py xp_fund_extractor.py --test 10     # testar com 10 fundos
    py xp_fund_extractor.py --start 50    # pular os primeiros 50 pendentes
"""

import os
import time
import json
import logging
import requests
import pandas as pd
import anthropic
import base64
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path

# ─── CONFIGURAÇÃO ────────────────────────────────────────────────────────────

INPUT_XP_FILE  = "Fundos_XP_raw.xlsx"       # Nome da sua planilha XP
OUTPUT_FILE    = "Modelo_preenchido.xlsx"    # Arquivo de saída
PDF_CACHE_DIR  = Path("pdfs_xp")            # Cache local dos PDFs
PROGRESS_FILE  = "progress_v2.json"         # Checkpoint de progresso

ANTHROPIC_API_KEY = "COLE_SUA_CHAVE_AQUI"   # Ex: sk-ant-api03-...
MODEL             = "claude-sonnet-4-20250514"

DELAY_BETWEEN_REQUESTS = 1.5  # segundos entre chamadas

# ─── COOKIES DE SESSÃO DA XP ─────────────────────────────────────────────────
# Como obter:
#   1. Acesse fundos.xpi.com.br no Edge (logado)
#   2. F12 → aba Rede → F5 → clique em qualquer requisição
#   3. Em "Cabeçalhos de solicitação", copie o valor completo da linha "cookie:"
#   4. Cole abaixo substituindo o texto entre as aspas triplas
# ATENÇÃO: cookies expiram em algumas horas.

SESSION_COOKIES = """COLE_AQUI_O_VALOR_COMPLETO_DA_LINHA_COOKIE_DO_EDGE"""

COLUNAS_MODELO = [
    "Nome", "CNPJ", "Categoria", "Subcategoria", "Gestor",
    "Indexador", "Liquidez", "Tributação", "Come-Cotas",
    "Taxa de Adm", "Taxa de Performance", "Benchmark",
    "Descrição Tributação", "Público Alvo", "Horizonte Mínimo (anos)",
    "Quando Indicar", "Quando não indicar", "Vantagens",
    "Desvantagens", "Alertas", "Descrição Simples", "Descrição Técnica",
]

# ─── LOGGING ─────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("extrator_xp.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ─── SYSTEM PROMPT ───────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Você é um especialista em fundos de investimento brasileiros com certificação CFP, \
atuando como analista editorial da 3A Riva Investimentos.

Sua função é analisar lâminas e regulamentos de fundos e produzir conteúdo objetivo, específico e \
acionável para assessores de investimento usarem no atendimento a clientes.

Escreva sempre com precisão técnica, mas sem exageros. Evite afirmações vagas como "boa gestão" ou \
"equipe experiente" sem embasamento concreto no documento. Nunca mencione rentabilidade passada como \
garantia de retorno futuro.

Responda SOMENTE com um objeto JSON válido, sem texto adicional, sem blocos de código markdown."""

# ─── PROMPT ANALÍTICO ────────────────────────────────────────────────────────

ANALYSIS_PROMPT = """Analise a lâmina/regulamento deste fundo e retorne um JSON com exatamente estas 7 chaves.

Contexto já conhecido sobre o fundo (use como base):
- Nome: {nome}
- Categoria XP: {categoria}
- Benchmark: {benchmark}
- Risco declarado: {risco}
- Liquidez (cotização + liquidação): {liquidez}
- Taxa de administração: {taxa_adm}% a.a.
- Taxa de performance: {taxa_perf}%
- Aplicação mínima: R$ {aplic_minima}
- Público-alvo: {publico}
- Rentabilidade 12m: {rent12m}% | 24m: {rent24m}% | 36m: {rent36m}%
- Patrimônio Líquido: R$ {pl}

Escala de horizonte: Curto (até 1 ano) | Médio (1 a 3 anos) | Longo (3+ anos)
Objetivos do investidor: Aposentadoria | Reserva de emergência | Compra de imóvel | \
Independência financeira / renda passiva | Morar no exterior | Crescimento patrimonial

JSON esperado:
{{
  "Gestor": "Nome completo da gestora extraído do documento. Máx 40 chars.",

  "Quando Indicar": "Para qual perfil de investidor e objetivo de vida este fundo é adequado. \
Cruze perfil de risco (conservador/moderado/arrojado) com horizonte temporal (Curto/Médio/Longo) \
e objetivos de vida acima. Seja específico — ex: 'Indicado para investidores moderados a arrojados \
com horizonte longo (3+ anos) buscando Crescimento patrimonial ou Independência financeira'. Máx 80 palavras.",

  "Quando não indicar": "Para qual perfil este fundo NÃO é adequado. Prioridade: (1) perfil de risco \
incompatível com drawdown/volatilidade (cite o número exato se disponível); (2) liquidez incompatível \
com necessidade do investidor; (3) valor mínimo alto demais para o cliente. Máx 60 palavras.",

  "Vantagens": "Liste 3 vantagens concretas deste fundo frente à média da sua categoria XP. \
Compare rentabilidade, volatilidade, liquidez ou características únicas — sempre com números \
quando disponíveis no documento. Não use afirmações genéricas. Formato: frases curtas. Máx 80 palavras.",

  "Desvantagens": "Liste 2-3 desvantagens reais — taxa de adm acima da média da categoria, \
liquidez longa, come-cotas, PL pequeno para a estratégia, etc. Compare com pares quando possível. \
Máx 60 palavras.",

  "Alertas": "Alertas na ordem de prioridade: (1) rentabilidade abaixo do benchmark nos últimos \
12/24/36 meses — cite os números exatos; (2) PL baixo para a categoria (abaixo de R$ 50M é sinal \
de atenção para maioria das categorias); (3) liquidez longa (D+30 ou mais); (4) alertas regulatórios \
específicos do fundo. Se nenhum alerta relevante: escreva 'Sem alertas relevantes'. Máx 80 palavras.",

  "Descrição Simples": "Explique o fundo em 2-3 frases para um investidor leigo. \
O que o fundo faz, qual é o objetivo e para quem serve. Sem jargão técnico. Máx 60 palavras.",

  "Descrição Técnica": "Descrição técnica completa: estratégia de investimento, principais ativos, \
benchmarks, limites de concentração relevantes, instrumentos utilizados (derivativos, crédito privado, \
etc.), diferenciais da gestão e principais riscos. Para uso do assessor. Máx 120 palavras."
}}"""

# ─── CHECKPOINT ──────────────────────────────────────────────────────────────

def carregar_checkpoint() -> dict:
    try:
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def salvar_checkpoint(checkpoint: dict) -> None:
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(checkpoint, f, ensure_ascii=False, indent=2)

# ─── PDF ─────────────────────────────────────────────────────────────────────

def baixar_pdf(url: str, nome: str, cookies_str: str) -> bytes | None:
    """Baixa PDF da URL com cookies de sessão XP."""
    PDF_CACHE_DIR.mkdir(exist_ok=True)
    nome_arquivo = "".join(c for c in nome if c.isalnum() or c in " _-")[:50] + ".pdf"
    caminho = PDF_CACHE_DIR / nome_arquivo

    if caminho.exists():
        log.info(f"  PDF em cache: {caminho}")
        return caminho.read_bytes()

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://fundos.xpi.com.br/",
        "Accept": "application/pdf,*/*",
    }
    cookies = {}
    for item in cookies_str.strip().split(";"):
        item = item.strip()
        if "=" in item:
            k, v = item.split("=", 1)
            cookies[k.strip()] = v.strip()

    try:
        resp = requests.get(url, headers=headers, cookies=cookies, timeout=30)
        if resp.status_code == 200 and len(resp.content) > 1000:
            caminho.write_bytes(resp.content)
            log.info(f"  PDF baixado: {len(resp.content)/1024:.0f} KB")
            return resp.content
        else:
            log.warning(f"  Falha no download: HTTP {resp.status_code}")
            return None
    except Exception as e:
        log.error(f"  Erro download: {e}")
        return None

# ─── ANÁLISE ─────────────────────────────────────────────────────────────────

def analisar_fundo(client: anthropic.Anthropic, fundo: dict, pdf_bytes: bytes) -> dict:
    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
    prompt = ANALYSIS_PROMPT.format(
        nome=fundo.get("product", ""),
        categoria=fundo.get("classificationXp", ""),
        benchmark=fundo.get("benchmark", ""),
        risco=fundo.get("risk", ""),
        liquidez=f"D+{fundo.get('quotationDays',0)} cotização / D+{fundo.get('liquidationDays',0)} liquidação",
        taxa_adm=fundo.get("managementFee", ""),
        taxa_perf=fundo.get("performanceFee", ""),
        aplic_minima=fundo.get("minimumInitialInvestment", ""),
        publico=fundo.get("targetAudience", ""),
        rent12m=fundo.get("return12m", ""),
        rent24m=fundo.get("return24m", ""),
        rent36m=fundo.get("return36m", ""),
        pl=fundo.get("netAssetValue", ""),
    )

    resp = client.messages.create(
        model=MODEL,
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                {"type": "text", "text": prompt},
            ]
        }]
    )
    texto = resp.content[0].text.strip()
    if "```" in texto:
        texto = texto.split("```")[1]
        if texto.startswith("json"):
            texto = texto[4:]
    return json.loads(texto)

# ─── SALVAR EXCEL ────────────────────────────────────────────────────────────

def salvar_excel(resultados: list, caminho: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fundos XP"

    hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Arial")

    for col_idx, col_name in enumerate(COLUNAS_MODELO, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    larguras = {
        "Nome": 35, "Categoria": 15, "Subcategoria": 30, "Gestor": 30,
        "Indexador": 12, "Liquidez": 25, "Tributação": 15, "Come-Cotas": 20,
        "Taxa de Adm": 12, "Taxa de Performance": 15, "Benchmark": 12,
        "Descrição Tributação": 45, "Público Alvo": 20, "Horizonte Mínimo (anos)": 20,
        "Quando Indicar": 50, "Quando não indicar": 50, "Vantagens": 50,
        "Desvantagens": 50, "Alertas": 50, "Descrição Simples": 60, "Descrição Técnica": 60,
    }
    for col_idx, col_name in enumerate(COLUNAS_MODELO, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 15)

    for row_idx, r in enumerate(resultados, 2):
        status = r.get("_status", "")
        fill = None
        if status == "erro":
            fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        elif status == "sem_pdf":
            fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
        for col_idx, col_name in enumerate(COLUNAS_MODELO, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(col_name, ""))
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(caminho)
    print(f"\n✅ Excel salvo: {caminho}")

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--test", type=int, default=0)
    parser.add_argument("--start", type=int, default=0)
    args = parser.parse_args()

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    df = pd.read_excel(INPUT_XP_FILE)
    fundos = df.to_dict("records")

    checkpoint = carregar_checkpoint()
    processados = set(checkpoint.keys())
    pendentes = [f for f in fundos if str(f.get("cnpj", "")) not in processados]

    if args.start:
        pendentes = pendentes[args.start:]
    if args.test:
        pendentes = pendentes[:args.test]

    print(f"📋 {len(pendentes)} fundos pendentes | {len(processados)} já processados")

    for i, fundo in enumerate(pendentes, 1):
        cnpj = str(fundo.get("cnpj", f"sem_cnpj_{i}"))
        nome = fundo.get("product", "?")
        print(f"\n[{i}/{len(pendentes)}] {nome[:60]}")

        resultado = {"Nome": nome, "CNPJ": cnpj, "_status": "ok"}
        try:
            # Campos diretos da planilha
            resultado["Categoria"] = fundo.get("classificationXp", "")
            resultado["Benchmark"] = fundo.get("benchmark", "")
            resultado["Tributação"] = fundo.get("taxationType", "")
            resultado["Come-Cotas"] = "Sim" if fundo.get("comeCota") else "Não"
            resultado["Taxa de Adm"] = fundo.get("managementFee", "")
            resultado["Taxa de Performance"] = fundo.get("performanceFee", "")
            resultado["Público Alvo"] = fundo.get("targetAudience", "")

            liquidez_cot = fundo.get("quotationDays", "")
            liquidez_liq = fundo.get("liquidationDays", "")
            resultado["Liquidez"] = f"D+{liquidez_cot} / D+{liquidez_liq}"

            # PDF
            pdf_url = fundo.get("linkLamina") or fundo.get("linkRegulamento")
            if not pdf_url:
                resultado["_status"] = "sem_pdf"
                resultado["_erro"] = "PDF não encontrado"
            else:
                pdf_bytes = baixar_pdf(pdf_url, nome, SESSION_COOKIES)
                if not pdf_bytes:
                    resultado["_status"] = "sem_pdf"
                    resultado["_erro"] = "Falha no download"
                else:
                    campos_ia = analisar_fundo(client, fundo, pdf_bytes)
                    resultado.update(campos_ia)

        except Exception as e:
            resultado["_status"] = "erro"
            resultado["_erro"] = str(e)
            log.error(f"  ERRO: {e}")

        checkpoint[cnpj] = resultado
        salvar_checkpoint(checkpoint)
        time.sleep(DELAY_BETWEEN_REQUESTS)

    # Gerar Excel com todos os processados
    todos = list(checkpoint.values())
    salvar_excel(todos, OUTPUT_FILE)
    ok = sum(1 for r in todos if r.get("_status") == "ok")
    erros = sum(1 for r in todos if r.get("_status") == "erro")
    sem_pdf = sum(1 for r in todos if r.get("_status") == "sem_pdf")
    print(f"\n📊 Total: {len(todos)} | ✅ {ok} | ⚠️ sem_pdf: {sem_pdf} | ❌ erros: {erros}")


if __name__ == "__main__":
    main()
