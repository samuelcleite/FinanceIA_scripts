"""
btg_fund_extractor.py
Pipeline FinanceIA — BTG Pactual
Lê fundos_btg_raw.json → preenche campos diretos → gera campos analíticos via Claude API (PDF via URL)
Output: fundos_btg_analisados.xlsx

Uso:
    python btg_fund_extractor.py --input fundos_btg_raw.json --output fundos_btg_analisados.xlsx
    python btg_fund_extractor.py --input fundos_btg_raw.json --output fundos_btg_analisados.xlsx --test 5
    python btg_fund_extractor.py --input fundos_btg_raw.json --output fundos_btg_analisados.xlsx --reset
"""

import argparse
import hashlib
import json
import os
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill

# ──────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ──────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = "COLE_SUA_CHAVE_AQUI"   # sk-ant-api03-...
CLAUDE_MODEL      = "claude-sonnet-4-20250514"
REQUEST_DELAY     = 1.5  # segundos entre chamadas

PDF_CACHE_DIR  = Path("pdfs_btg")
CHECKPOINT_FILE = Path("progress_btg.json")

# Prioridade de PDF: Informativo Mensal → Carta da Gestão → Lâmina → Regulamento
PRIORIDADE_PDF = [
    "Informativo Mensal",
    "Carta da Gestão",
    "Lâmina de Informações Essenciais",
    "Regulamento de Fundo",
]

# Mapeamento de classe BTG → Categoria do Modelo
CLASSE_TO_CATEGORIA = {
    "Renda Fixa":                   "Renda Fixa",
    "Renda Fixa Crédito Privado":   "Renda Fixa",
    "Referenciado DI":              "Renda Fixa",
    "Multimercado":                 "Multimercado",
    "Ações":                        "Renda Variável",
    "Long Biased":                  "Renda Variável",
    "Long Only":                    "Renda Variável",
    "Fundos Imobiliários":          "Fundos Imobiliários",
    "Previdência Renda Fixa":       "Previdência",
    "Previdência Multimercado":     "Previdência",
    "Previdência Ações":            "Previdência",
    "RF Global":                    "Internacional",
    "RV Global":                    "Internacional",
    "Commodities e Moedas":         "Internacional",
    "Cambial":                      "Cambial",
}

COLUNAS_MODELO = [
    "Nome", "CNPJ", "Categoria", "Subcategoria", "Gestor",
    "Indexador", "Liquidez", "Tributação", "Come-Cotas",
    "Taxa de Adm", "Taxa de Performance", "Benchmark",
    "Descrição Tributação", "Público Alvo", "Horizonte Mínimo (anos)",
    "Quando Indicar", "Quando não indicar", "Vantagens",
    "Desvantagens", "Alertas", "Descrição Simples", "Descrição Técnica",
]

COLUNAS_BTG_EXTRA = [
    "_volatilidade_12m", "_sharpe_12m", "_retorno_12m",
    "_retorno_24m", "_retorno_36m", "_pl", "_fonte_pdf", "_status", "_erro",
]

ALL_COLUMNS = COLUNAS_MODELO + COLUNAS_BTG_EXTRA

# ──────────────────────────────────────────────────────────────
# PROMPTS
# ──────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Você é um especialista em fundos de investimento brasileiros com certificação CFP, \
atuando como analista editorial da 3A Riva Investimentos.

Sua função é analisar lâminas e regulamentos de fundos e produzir conteúdo objetivo, específico e \
acionável para assessores de investimento usarem no atendimento a clientes.

Escreva sempre com precisão técnica, mas sem exageros. Evite afirmações vagas como "boa gestão" ou \
"equipe experiente" sem embasamento concreto no documento. Nunca mencione rentabilidade passada como \
garantia de retorno futuro.

Responda SOMENTE com um objeto JSON válido, sem texto adicional, sem blocos de código markdown."""

ANALYSIS_PROMPT_WITH_PDF = """Analise o documento PDF deste fundo e o contexto abaixo. \
Retorne um JSON com exatamente estas 8 chaves.

Contexto do fundo (dados estruturados do BTG):
- Nome: {nome}
- Gestor: {gestor}
- Categoria BTG: {categoria_btg}
- Subcategoria: {subcategoria}
- Benchmark: {benchmark}
- Taxa de Adm: {taxa_adm}% a.a.
- Taxa de Performance: {taxa_perf}%
- Volatilidade 12m: {volatilidade_12m}% a.a.
- Sharpe 12m: {sharpe_12m}
- Retorno 12m: {retorno_12m}% | 24m: {retorno_24m}% | 36m: {retorno_36m}%
- PL: R$ {pl}
- Liquidez: {liquidez}
- Aplicação Mínima: R$ {aplic_minima}

Escala de horizonte: Curto (até 1 ano) | Médio (1 a 3 anos) | Longo (3+ anos)
Objetivos: Aposentadoria | Reserva de emergência | Compra de imóvel | Independência financeira | Crescimento patrimonial

JSON:
{{
  "Gestor": "Nome completo da gestora conforme o documento.",
  "Horizonte Mínimo (anos)": "Curto (até 1 ano)" | "Médio (1 a 3 anos)" | "Longo (3+ anos)",
  "Quando Indicar": "Para qual perfil e objetivo. Cruze risco × horizonte × objetivo. Máx 80 palavras.",
  "Quando não indicar": "Perfis/situações onde o fundo não é adequado. Cite volatilidade exata. Máx 60 palavras.",
  "Vantagens": "3 vantagens concretas vs média da categoria BTG. Use números. Máx 80 palavras.",
  "Desvantagens": "2-3 desvantagens reais (taxa, liquidez, PL, come-cotas). Máx 60 palavras.",
  "Alertas": "Prioridade: (1) rentabilidade abaixo do benchmark — cite números; (2) PL baixo; (3) liquidez longa; (4) regulatórios. Máx 80 palavras.",
  "Descrição Simples": "2-3 frases para o cliente leigo. O que faz, objetivo, para quem. Máx 60 palavras.",
  "Descrição Técnica": "Estratégia, ativos, benchmarks, instrumentos, diferenciais, riscos. Para o assessor. Máx 120 palavras."
}}"""

ANALYSIS_PROMPT_SEM_PDF = """Com base APENAS nos dados estruturados abaixo (sem PDF disponível), \
retorne um JSON com exatamente estas 8 chaves.

Dados do fundo:
- Nome: {nome}
- Gestor: {gestor}
- Categoria BTG: {categoria_btg}
- Subcategoria: {subcategoria}
- Benchmark: {benchmark}
- Taxa de Adm: {taxa_adm}% a.a.
- Taxa de Performance: {taxa_perf}%
- Volatilidade 12m: {volatilidade_12m}% a.a.
- Sharpe 12m: {sharpe_12m}
- Retorno 12m: {retorno_12m}% | 24m: {retorno_24m}% | 36m: {retorno_36m}%
- Liquidez: {liquidez}

Baseie-se apenas no contexto acima.

JSON:
{{
  "Gestor": "string",
  "Horizonte Mínimo (anos)": "string",
  "Quando Indicar": "string",
  "Quando não indicar": "string",
  "Vantagens": "string",
  "Desvantagens": "string",
  "Alertas": "string",
  "Descrição Simples": "string",
  "Descrição Técnica": "string"
}}"""

# ──────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────

def _parse_json(text: str) -> dict:
    text = text.strip()
    if "```" in text:
        parts = text.split("```")
        for p in parts:
            p = p.strip()
            if p.startswith("json"):
                p = p[4:]
            if p.strip().startswith("{"):
                text = p.strip()
                break
    return json.loads(text)


def extrair_campos_diretos(fundo: dict) -> dict:
    nome      = fundo.get("product", "")
    cnpj      = fundo.get("cnpj", "")
    classe    = fundo.get("class", "")
    gestor    = fundo.get("manager", "")
    benchmark = fundo.get("benchmark", "")

    categoria = CLASSE_TO_CATEGORIA.get(classe, "Outros")

    cotiz = fundo.get("quotationDays", "")
    liq   = fundo.get("liquidationDays", "")
    liquidez_str = f"D+{cotiz} cotização / D+{liq} liquidação" if cotiz else ""

    come_cotas = "Sim" if fundo.get("comeCotas") else "Não"
    tributacao = fundo.get("taxation", "")

    desc_trib_map = {
        "Longo Prazo": "IR regressivo: 22,5% (até 180 dias) a 15% (acima de 720 dias). Come-cotas semestral.",
        "Curto Prazo": "IR regressivo: 22,5% (até 180 dias) a 20% (acima). Come-cotas semestral em maio e novembro.",
        "Ações":       "IR 15% sobre ganhos. Sem come-cotas. IOF nos primeiros 30 dias.",
        "Exclusivo":   "Tributação exclusiva conforme regulamento.",
    }
    descricao_tributacao = desc_trib_map.get(tributacao, "")

    taxa_adm  = fundo.get("managementFee", "")
    taxa_perf = fundo.get("performanceFee", "")
    aplic_min = fundo.get("minimumInitialInvestment", "")
    publico   = fundo.get("targetAudience", "")
    subcategoria = fundo.get("subclass", "")

    # Extras para referência
    extras = {
        "_volatilidade_12m": fundo.get("annualizedVolatility12m", ""),
        "_sharpe_12m":       fundo.get("sharpeRatio12m", ""),
        "_retorno_12m":      fundo.get("return12m", ""),
        "_retorno_24m":      fundo.get("return24m", ""),
        "_retorno_36m":      fundo.get("return36m", ""),
        "_pl":               fundo.get("netAssetValue", ""),
    }

    campos = {
        "Nome":                  nome,
        "CNPJ":                  cnpj,
        "Categoria":             categoria,
        "Subcategoria":          subcategoria,
        "Gestor":                gestor,
        "Indexador":             benchmark,
        "Liquidez":              liquidez_str,
        "Tributação":            tributacao,
        "Come-Cotas":            come_cotas,
        "Taxa de Adm":           f"{taxa_adm}%" if taxa_adm else "",
        "Taxa de Performance":   f"{taxa_perf}%" if taxa_perf else "",
        "Benchmark":             benchmark,
        "Descrição Tributação":  descricao_tributacao,
        "Público Alvo":          publico,
        "Horizonte Mínimo (anos)": "",
    }
    campos.update(extras)
    return campos


def montar_contexto(fundo: dict, campos: dict) -> dict:
    return {
        "nome":           campos.get("Nome", ""),
        "gestor":         campos.get("Gestor", ""),
        "categoria_btg":  fundo.get("class", ""),
        "subcategoria":   campos.get("Subcategoria", ""),
        "benchmark":      campos.get("Benchmark", ""),
        "taxa_adm":       fundo.get("managementFee", ""),
        "taxa_perf":      fundo.get("performanceFee", ""),
        "volatilidade_12m": campos.get("_volatilidade_12m", ""),
        "sharpe_12m":     campos.get("_sharpe_12m", ""),
        "retorno_12m":    campos.get("_retorno_12m", ""),
        "retorno_24m":    campos.get("_retorno_24m", ""),
        "retorno_36m":    campos.get("_retorno_36m", ""),
        "pl":             campos.get("_pl", ""),
        "liquidez":       campos.get("Liquidez", ""),
        "aplic_minima":   fundo.get("minimumInitialInvestment", ""),
    }


def selecionar_pdf(fundo: dict):
    files = fundo.get("detail", {}).get("files", []) or []
    by_desc = {f["description"]: f for f in files if f.get("url")}
    for prioridade in PRIORIDADE_PDF:
        if prioridade in by_desc:
            return by_desc[prioridade]["url"], prioridade
    return None, None


def baixar_pdf(url: str, cnpj: str) -> bytes | None:
    PDF_CACHE_DIR.mkdir(exist_ok=True)
    url_hash = hashlib.md5(url.encode()).hexdigest()[:10]
    cache_path = PDF_CACHE_DIR / f"{cnpj}_{url_hash}.pdf"
    if cache_path.exists():
        return cache_path.read_bytes()

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/pdf,*/*",
        "Referer": "https://portal.btgpactual.com/",
    }
    for tentativa in range(3):
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code == 200 and len(resp.content) > 1000:
                cache_path.write_bytes(resp.content)
                return resp.content
            print(f"    ⚠️  Status {resp.status_code} (tentativa {tentativa+1})")
        except Exception as e:
            print(f"    ⚠️  Erro (tentativa {tentativa+1}): {e}")
        if tentativa < 2:
            time.sleep(2 ** tentativa)
    return None


# ──────────────────────────────────────────────────────────────
# CHAMADA CLAUDE API
# ──────────────────────────────────────────────────────────────

def chamar_claude_com_pdf(client: anthropic.Anthropic, contexto: dict, pdf_url: str) -> dict:
    """PDF entregue via URL — evita timeout por payload grande."""
    prompt = ANALYSIS_PROMPT_WITH_PDF.format(**contexto)
    resp = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=2000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document", "source": {"type": "url", "url": pdf_url}},
                {"type": "text", "text": prompt},
            ]
        }]
    )
    return _parse_json(resp.content[0].text)


def chamar_claude_sem_pdf(client: anthropic.Anthropic, contexto: dict) -> dict:
    prompt = ANALYSIS_PROMPT_SEM_PDF.format(**contexto)
    resp = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=1500,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}],
    )
    return _parse_json(resp.content[0].text)


# ──────────────────────────────────────────────────────────────
# CHECKPOINT
# ──────────────────────────────────────────────────────────────

def carregar_checkpoint() -> dict:
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def salvar_checkpoint(checkpoint: dict) -> None:
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(checkpoint, f, ensure_ascii=False, indent=2)


# ──────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────

def salvar_excel(rows: list, output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fundos BTG"

    hdr_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10, name="Arial")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align

    ws.row_dimensions[1].height = 30

    fill_erro   = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    fill_sempdf = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    body_font   = Font(name="Arial", size=10)

    for row_idx, row in enumerate(rows, 2):
        status = row.get("_status", "")
        fill   = fill_erro if status == "erro" else (fill_sempdf if status == "sem_pdf" else None)
        for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.font      = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    larguras = {
        "Nome": 38, "CNPJ": 16, "Categoria": 15, "Subcategoria": 30, "Gestor": 30,
        "Indexador": 12, "Liquidez": 28, "Tributação": 15, "Come-Cotas": 12,
        "Taxa de Adm": 12, "Taxa de Performance": 18, "Benchmark": 15,
        "Descrição Tributação": 45, "Público Alvo": 25, "Horizonte Mínimo (anos)": 20,
        "Quando Indicar": 50, "Quando não indicar": 50, "Vantagens": 50,
        "Desvantagens": 50, "Alertas": 50, "Descrição Simples": 60, "Descrição Técnica": 60,
    }
    for col_idx, col_name in enumerate(ALL_COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 14)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\n✅ Excel salvo: {output_path}")


# ──────────────────────────────────────────────────────────────
# PIPELINE PRINCIPAL
# ──────────────────────────────────────────────────────────────

def processar_fundos(fundos, client, checkpoint, modo_teste=0, start_from=0):
    resultados = list(checkpoint.values())
    processados = set(checkpoint.keys())
    pendentes = [f for f in fundos if f.get("cnpj") not in processados]

    if start_from:
        pendentes = pendentes[start_from:]
    if modo_teste:
        pendentes = pendentes[:modo_teste]

    total = len(pendentes)
    print(f"\n📋 {total} pendentes | {len(processados)} já processados\n")

    for i, fundo in enumerate(pendentes, 1):
        nome = fundo.get("product", "?")
        cnpj = fundo.get("cnpj", f"sem_cnpj_{i}")
        print(f"[{i:3d}/{total}] {nome[:60]}")

        resultado = {}
        try:
            campos_diretos = extrair_campos_diretos(fundo)
            resultado.update(campos_diretos)
            pdf_url, pdf_desc = selecionar_pdf(fundo)
            contexto = montar_contexto(fundo, campos_diretos)

            if pdf_url:
                print(f"       📄 PDF: {pdf_desc}")
                campos_ia = chamar_claude_com_pdf(client, contexto, pdf_url)
                resultado.update(campos_ia)
                resultado["_fonte_pdf"] = pdf_desc
                resultado["_status"] = "ok"
                print(f"       ✅ OK")
            else:
                print(f"       ⚠️  Sem PDF — usando metadados")
                campos_ia = chamar_claude_sem_pdf(client, contexto)
                resultado.update(campos_ia)
                resultado["_fonte_pdf"] = "sem_pdf"
                resultado["_status"] = "sem_pdf"

            resultado["_erro"] = ""
        except Exception as e:
            print(f"       ❌ ERRO: {e}")
            resultado["_status"] = "erro"
            resultado["_erro"] = str(e)
            traceback.print_exc()

        resultados.append(resultado)
        checkpoint[cnpj] = resultado
        salvar_checkpoint(checkpoint)
        time.sleep(REQUEST_DELAY)

    return resultados


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",  required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--test",   type=int, default=0)
    parser.add_argument("--start",  type=int, default=0)
    parser.add_argument("--reset",  action="store_true")
    args = parser.parse_args()

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    with open(args.input, encoding="utf-8") as f:
        fundos = json.load(f)
    print(f"📂 {len(fundos)} fundos carregados")

    checkpoint = {} if args.reset else carregar_checkpoint()

    inicio = datetime.now()
    resultados = processar_fundos(fundos, client, checkpoint, args.test, args.start)
    fim = datetime.now()

    salvar_excel(resultados, args.output)

    ok    = sum(1 for r in resultados if r.get("_status") == "ok")
    erros = sum(1 for r in resultados if r.get("_status") == "erro")
    sem   = sum(1 for r in resultados if r.get("_status") == "sem_pdf")
    duracao = (fim - inicio).total_seconds()
    print(f"\n📊 Total: {len(resultados)} | ✅ {ok} | ⚠️ sem_pdf: {sem} | ❌ erros: {erros}")
    print(f"⏱️  {duracao:.0f}s ({duracao/60:.1f} min)")


if __name__ == "__main__":
    main()
